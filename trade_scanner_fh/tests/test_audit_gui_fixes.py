"""GUI-side regression tests for audit fixes (H1, M7, L6, L9, L10)."""
from __future__ import annotations

from unittest.mock import MagicMock

import pandas as pd
import pytest


@pytest.fixture(autouse=True)
def _no_launch_data_pipeline(monkeypatch):
    """The full-``MainWindow()`` tests below construct a real window to
    exercise widget/column/export behavior — they don't test the launch-
    time data pipeline. ``MainWindow.__init__`` starts a UniverseRefresh
    worker whose ``finished`` signal calls ``_load_universe_and_update``,
    which (when the OHLCV cache is current) now kicks off the earnings
    smart refresh's real network fill workers. Spawning those during
    pytest-qt event processing corrupts the heap (Windows 0xc0000374).
    Stub the entry point so no launch-time fills run; these tests seed
    their own results directly."""
    from trade_scanner_fh.gui.main_window import MainWindow
    monkeypatch.setattr(
        MainWindow, "_load_universe_and_update", lambda self: None,
    )


# ──────────────────────────────────────────────────────────────────────
# L6 — green-highlight detection uses explicit suffix prefix, not "in key"
# ──────────────────────────────────────────────────────────────────────

def test_l6_green_highlight_only_on_metric_specific_columns(_qapp):
    """A future hypothetical column with both 'rev' and 'eps' in the
    name (or some other ambiguous substring) must NOT silently pick up
    streak coloring. The new check uses _Q_COL_RE + suffix prefix
    matching, so unrelated columns are immune."""
    from trade_scanner_fh.gui.widgets import ResultsTable

    # Include `reported_eps` so the "Curr Reported EPS" column actually
    # appears in the table — the 2026-05 update filters columns whose
    # key isn't present in the df.
    df = pd.DataFrame([{
        "symbol": "A", "close": 100.0, "price": 100.0, "pct_gain": 10.0,
        "reported_eps": 2.5,
        "consec_eps_beats": 2,
        "q1_reported_eps": 2.0, "q1_surprise_eps_dollar": 0.1,
        "q1_surprise_eps_pct": 5.0,
        "q2_reported_eps": 1.9, "q2_surprise_eps_dollar": 0.05,
        "q2_surprise_eps_pct": 3.0,
    }])

    table = ResultsTable()
    table.populate(df)

    cols = table.active_columns
    headers = [c[0] for c in cols]
    # Q-1 is in-streak (eps_streak=2 ≥ 1)
    q1_idx = headers.index("Q-1 Reported EPS")
    item = table.model_src.item(0, q1_idx)
    assert item.foreground().color() == ResultsTable._STREAK_GREEN

    # Non-q* column "Curr Reported EPS" is NOT a streak cell; default color.
    # The header was renamed from "Reported EPS" to "Curr Reported EPS"
    # alongside the filter labels gaining a "Current" prefix.
    rep_eps_idx = headers.index("Curr Reported EPS")
    rep_item = table.model_src.item(0, rep_eps_idx)
    assert rep_item.foreground().color() != ResultsTable._STREAK_GREEN


# ──────────────────────────────────────────────────────────────────────
# L10 — _Q_COL_RE precompiled regex matches q-prefix columns
# ──────────────────────────────────────────────────────────────────────

def test_l10_q_col_re_extracts_quarter_index_and_suffix():
    from trade_scanner_fh.gui.widgets import _Q_COL_RE

    m = _Q_COL_RE.match("q12_surprise_eps_dollar")
    assert m is not None
    assert m.group(1) == "12"
    assert m.group(2) == "surprise_eps_dollar"

    # Non-q-prefix shouldn't match
    assert _Q_COL_RE.match("reported_eps") is None
    assert _Q_COL_RE.match("question_mark") is None


# ──────────────────────────────────────────────────────────────────────
# L9 — ZacksFillWorker exception path surfaces partial candidates
# ──────────────────────────────────────────────────────────────────────

def test_l9_inner_fill_exception_emits_partial_candidates(_qapp, tmp_path,
                                                           monkeypatch):
    """When the inner fill raises mid-flight, finished() must surface
    the in-flight `candidates` list so a slot can offer retry options."""
    from trade_scanner_fh import config as cfg
    from trade_scanner_fh import earnings_history as eh
    monkeypatch.setattr(cfg, "DATA_DIR", tmp_path)
    monkeypatch.setattr(cfg, "EARNINGS_HISTORY_PARQUET",
                        tmp_path / "earnings_history.parquet")
    (tmp_path / "scanner_data").mkdir(parents=True, exist_ok=True)

    from trade_scanner_fh.gui.workers import ZacksFillWorker

    # Make has_zacks_cookies True so the worker attempts the fill.
    from trade_scanner_fh import zacks_scraper as zs
    monkeypatch.setattr(zs, "get_zacks_cookies", lambda: "k=v")

    # Force the fill function to raise.
    def boom(*a, **kw):
        raise RuntimeError("simulated inner failure")

    monkeypatch.setattr(eh, "targeted_fill_zacks", boom)
    monkeypatch.setattr(eh, "bulk_fill_zacks", boom)

    captured = {}

    worker = ZacksFillWorker(
        ["AAA", "BBB", "CCC"], blacklist=set(), mode="targeted",
    )
    worker.finished.connect(
        lambda f, e, c: captured.update(filled=f, errors=e, candidates=c)
    )
    # Run synchronously instead of starting a thread
    worker.run()

    assert captured["filled"] == 0
    assert captured["errors"] == 0
    # Partial candidates surfaced (audit L9): the worker captured the
    # candidate list before the inner fill blew up.
    assert captured["candidates"] == ["AAA", "BBB", "CCC"]


# ──────────────────────────────────────────────────────────────────────
# H1 / M7 — Both Imperva dialogs can be constructed (smoke test)
# ──────────────────────────────────────────────────────────────────────

def test_h1_smart_refresh_connects_imperva_signal(_qapp, monkeypatch, tmp_path):
    """Audit H1: the smart-refresh launch path must connect
    imperva_block_detected to the cookie-refresh slot — a bug that left
    the worker hanging forever on the resume Event when blocked."""
    # Verify by inspecting the slot connection list. We do this via the
    # source rather than a live launch (which depends on universe data).
    src = (
        __import__("trade_scanner_fh.gui.main_window", fromlist=["x"])
        .__file__
    )
    text = open(src, encoding="utf-8").read()
    # Find the _kick_off_zacks_smart_refresh function
    start = text.index("def _kick_off_zacks_smart_refresh")
    end = text.index("\n    def ", start + 1)
    body = text[start:end]
    assert "imperva_block_detected.connect" in body, (
        "smart-refresh path must wire the Imperva auto-pause dialog"
    )


def test_m7_cookie_textedit_helper_starts_masked_when_existing(_qapp):
    """The helper used by both cookie dialogs masks the textarea when
    seeded with existing content. Uses `__new__` to bypass MainWindow's
    full __init__ — the helper is a pure method and doesn't depend on
    instance state."""
    from trade_scanner_fh.gui.main_window import MainWindow

    helper = MainWindow._build_cookie_textedit
    instance = MainWindow.__new__(MainWindow)  # skip __init__
    # Hide-by-default for non-empty content
    txt, btn = helper(instance, "secret-cookie-string", hide_initially=True)
    assert btn.text() == "Show"
    # When initial empty, button reads Hide (visible by default)
    txt2, btn2 = helper(instance, "", hide_initially=True)
    assert btn2.text() == "Hide"


# ──────────────────────────────────────────────────────────────────────
# Send Misses → Zacks Skip List button (failures dialog)
# ──────────────────────────────────────────────────────────────────────

def _make_minimal_main_window_for_skip_list(tmp_path, monkeypatch):
    """Build a MainWindow shell that has just enough state to exercise
    `_send_zacks_misses_to_skip_list` — bypasses __init__ and wires
    the in-memory skip-list set + the on-disk path to a tmp file so
    the test doesn't touch the user's real scanner_data."""
    from trade_scanner_fh.gui.main_window import MainWindow
    from trade_scanner_fh import config

    monkeypatch.setattr(config, "DATA_DIR", tmp_path)

    parent = MainWindow.__new__(MainWindow)
    parent._zacks_blacklist = set()
    parent._ZACKS_BLACKLIST_FILE = tmp_path / "zacks_blacklist.txt"
    parent.log_panel = MagicMock()
    return parent


def test_send_misses_to_skip_list_dedupes_against_existing(_qapp, tmp_path, monkeypatch):
    """Pre-existing tickers on the skip list are NOT added again. The
    new-add count reflects only tickers that weren't already present."""
    from PyQt6.QtWidgets import QMessageBox
    from unittest.mock import patch

    parent = _make_minimal_main_window_for_skip_list(tmp_path, monkeypatch)
    parent._zacks_blacklist = {"AAPL", "MSFT"}  # pre-existing entries

    breakdown = {
        "blocked": ["AAPL", "GOOGL"],     # AAPL is dup, GOOGL is new
        "not_found": ["MSFT", "NVDA"],    # MSFT is dup, NVDA is new
        "http_error": ["TSLA"],           # TSLA is new
    }

    with patch.object(QMessageBox, "information") as info:
        parent._send_zacks_misses_to_skip_list(breakdown, parent_dlg=None)

    # 3 new (GOOGL, NVDA, TSLA), 2 already on list (AAPL, MSFT)
    assert "AAPL" in parent._zacks_blacklist
    assert "MSFT" in parent._zacks_blacklist
    assert "GOOGL" in parent._zacks_blacklist
    assert "NVDA" in parent._zacks_blacklist
    assert "TSLA" in parent._zacks_blacklist
    assert len(parent._zacks_blacklist) == 5

    # Dialog message reports the counts
    args = info.call_args
    msg = args[0][2] if len(args[0]) >= 3 else str(args)
    assert "<b>3</b>" in msg, f"expected 3 new, msg = {msg}"
    assert "<b>2</b>" in msg, f"expected 2 already-on-list, msg = {msg}"

    # File written: one ticker per line, sorted
    written = (tmp_path / "zacks_blacklist.txt").read_text(encoding="utf-8")
    lines = [ln for ln in written.split("\n") if ln.strip()]
    assert lines == ["AAPL", "GOOGL", "MSFT", "NVDA", "TSLA"]


def test_send_misses_to_skip_list_normalizes_unicode_dashes(_qapp, tmp_path, monkeypatch):
    """Unicode minus / en-dash / em-dash variants in ticker text get
    normalized to ASCII hyphen via `_normalize_ticker`."""
    from PyQt6.QtWidgets import QMessageBox
    from unittest.mock import patch

    parent = _make_minimal_main_window_for_skip_list(tmp_path, monkeypatch)
    breakdown = {
        "blocked": ["BRK—A", "BF–B", "RDS−A"],  # em / en / minus
    }
    with patch.object(QMessageBox, "information"):
        parent._send_zacks_misses_to_skip_list(breakdown, parent_dlg=None)

    assert parent._zacks_blacklist == {"BRK-A", "BF-B", "RDS-A"}


def test_send_misses_to_skip_list_idempotent_re_click_is_noop(_qapp, tmp_path, monkeypatch):
    """Re-clicking the button after a successful save is a no-op —
    everything is already on the list, no spurious double-add."""
    from PyQt6.QtWidgets import QMessageBox
    from unittest.mock import patch

    parent = _make_minimal_main_window_for_skip_list(tmp_path, monkeypatch)
    breakdown = {"blocked": ["AAPL", "MSFT"]}

    with patch.object(QMessageBox, "information"):
        parent._send_zacks_misses_to_skip_list(breakdown, parent_dlg=None)
    snapshot = set(parent._zacks_blacklist)
    assert snapshot == {"AAPL", "MSFT"}

    with patch.object(QMessageBox, "information") as info2:
        parent._send_zacks_misses_to_skip_list(breakdown, parent_dlg=None)
    assert parent._zacks_blacklist == snapshot
    title = info2.call_args[0][1]
    assert title == "Nothing New"


def test_send_misses_to_skip_list_empty_breakdown_shows_info(_qapp, tmp_path, monkeypatch):
    """Empty breakdown shows an info dialog rather than no-op."""
    from PyQt6.QtWidgets import QMessageBox
    from unittest.mock import patch

    parent = _make_minimal_main_window_for_skip_list(tmp_path, monkeypatch)

    with patch.object(QMessageBox, "information") as info:
        parent._send_zacks_misses_to_skip_list({}, parent_dlg=None)
    assert parent._zacks_blacklist == set()
    info.assert_called_once()
    title = info.call_args[0][1]
    assert title == "Nothing to Send"


def test_send_misses_to_skip_list_filters_blank_tickers(_qapp, tmp_path, monkeypatch):
    """Blank / whitespace-only tickers are skipped — they'd produce
    empty lines in the file otherwise."""
    from PyQt6.QtWidgets import QMessageBox
    from unittest.mock import patch

    parent = _make_minimal_main_window_for_skip_list(tmp_path, monkeypatch)
    breakdown = {"blocked": ["AAPL", "", "  ", "MSFT"]}
    with patch.object(QMessageBox, "information"):
        parent._send_zacks_misses_to_skip_list(breakdown, parent_dlg=None)
    assert parent._zacks_blacklist == {"AAPL", "MSFT"}


def test_send_misses_to_skip_list_save_failure_rolls_back(_qapp, tmp_path, monkeypatch):
    """If the atomic write fails, the in-memory set is rolled back so
    a retry sees the same state. User gets a critical-error dialog."""
    from PyQt6.QtWidgets import QMessageBox
    from unittest.mock import patch

    parent = _make_minimal_main_window_for_skip_list(tmp_path, monkeypatch)
    breakdown = {"blocked": ["AAPL", "MSFT"]}

    parent._save_zacks_blacklist = lambda: (_ for _ in ()).throw(
        OSError("simulated disk-full"),
    )

    with patch.object(QMessageBox, "critical") as crit, \
         patch.object(QMessageBox, "information"):
        parent._send_zacks_misses_to_skip_list(breakdown, parent_dlg=None)

    # Rollback: in-memory set is empty again
    assert parent._zacks_blacklist == set()
    crit.assert_called_once()
    title = crit.call_args[0][1]
    assert title == "Save Failed"


# ──────────────────────────────────────────────────────────────────────
# Surge mode-dependent grey-out — visual styling on the dark theme
# ──────────────────────────────────────────────────────────────────────

def test_surge_trend_mode_greys_days_with_explicit_style(_qapp):
    """Selecting Trend-Continuous as the surge mode must disable the
    `days` spinbox (no fixed window in trend mode) AND apply the
    explicit muted stylesheet (Qt's default `:disabled` look on a dark
    theme is too subtle to see). Symmetric: switching away from trend
    must clear the muted style and re-enable days."""
    from trade_scanner_fh.gui.widgets import IndicatorPanel, IndicatorRow

    panel = IndicatorPanel()
    row = panel.rows["surge"]
    combo = row.spinboxes["mode"]
    days = row.spinboxes["days"]
    max_dd = row.spinboxes["max_dd"]
    muted = IndicatorRow._GREYED_INPUT_STYLE

    # Default mode is trend → days disabled + muted, max_dd enabled + clean
    assert combo.currentData() == "trend"
    assert days.isEnabled() is False
    assert days.styleSheet() == muted, (
        f"days should have muted style under trend, got {days.styleSheet()!r}"
    )
    assert max_dd.isEnabled() is True
    assert max_dd.styleSheet() == ""

    # Switch to close-to-close → flip
    for i in range(combo.count()):
        if combo.itemData(i) == "close":
            combo.setCurrentIndex(i)
            break
    assert days.isEnabled() is True
    assert days.styleSheet() == "", (
        "days should clear muted style when re-enabled"
    )
    assert max_dd.isEnabled() is False
    assert max_dd.styleSheet() == muted

    # Switch back to trend → re-mute days
    for i in range(combo.count()):
        if combo.itemData(i) == "trend":
            combo.setCurrentIndex(i)
            break
    assert days.isEnabled() is False
    assert days.styleSheet() == muted
    assert max_dd.isEnabled() is True
    assert max_dd.styleSheet() == ""


def test_surge_greyout_resyncs_after_preset_load_into_trend(_qapp):
    """Loading a preset that lands the surge mode on 'trend' must
    re-apply the muted styling on `days` (the resync block in
    `from_dict` covers this — without it, the visual would be wrong
    after preset load even though the underlying enabled state was
    correct)."""
    from trade_scanner_fh.gui.widgets import IndicatorPanel, IndicatorRow

    panel = IndicatorPanel()
    row = panel.rows["surge"]
    days = row.spinboxes["days"]
    muted = IndicatorRow._GREYED_INPUT_STYLE

    # Switch to a non-trend mode first so we have something to flip from
    combo = row.spinboxes["mode"]
    for i in range(combo.count()):
        if combo.itemData(i) == "close":
            combo.setCurrentIndex(i)
            break
    assert days.styleSheet() == ""

    # Now load a preset that sets mode back to trend
    panel.from_dict({"surge": {"enabled": True, "mode": "trend",
                               "min_pct": 40.0, "days": 7,
                               "max_dd": 25.0}})
    # days must be muted again
    assert days.isEnabled() is False
    assert days.styleSheet() == muted, (
        f"after preset load to trend, days should be muted; got {days.styleSheet()!r}"
    )


# ──────────────────────────────────────────────────────────────────────
# ResultsTable.populate re-entrancy guard (timeframe dropdown hang fix)
# ──────────────────────────────────────────────────────────────────────

def test_populate_drops_re_entrant_call(_qapp):
    """populate() must drop a re-entrant call (made from inside
    processEvents() during the chunked render) rather than letting
    the inner call corrupt the outer one's model state. Without
    this guard, a rapid second timeframe-dropdown click during a
    big render would re-enter `_on_timeframe_changed` → populate(),
    `setRowCount(0)` would clear the model the outer populate is
    still writing to, and the user would see a hang (stack-deep
    recursion or model corruption)."""
    import pandas as pd
    from trade_scanner_fh.gui.widgets import ResultsTable

    table = ResultsTable()
    df = pd.DataFrame([
        {"symbol": f"T{i:04d}", "close": 100.0 + i, "price": 100.0 + i,
         "pct_gain": 1.0 + i * 0.01}
        for i in range(50)
    ])

    call_count = {"n": 0}
    orig_impl = table._populate_impl

    def re_entrant_impl(d):
        call_count["n"] += 1
        if call_count["n"] == 1:
            # While "rendering" the first call, simulate a re-entrant
            # populate from another path (e.g. dropdown click).
            inner_df = pd.DataFrame([{"symbol": "X", "close": 1.0,
                                      "price": 1.0, "pct_gain": 0.0}])
            table.populate(inner_df)  # ← MUST be dropped, not recurse
        orig_impl(d)

    table._populate_impl = re_entrant_impl
    table.populate(df)
    # Outer populate ran exactly once; inner re-entrant call was dropped
    # before reaching _populate_impl.
    assert call_count["n"] == 1
    # Outer's data made it into the model — wasn't clobbered by inner
    assert table.model_src.rowCount() == 50


def test_populate_in_flight_flag_clears_after_normal_run(_qapp):
    """The `_populate_in_flight` flag must clear in the `finally`
    block so a second NORMAL (non-re-entrant) populate works."""
    import pandas as pd
    from trade_scanner_fh.gui.widgets import ResultsTable

    table = ResultsTable()
    df1 = pd.DataFrame([
        {"symbol": "A", "close": 100.0, "price": 100.0, "pct_gain": 5.0},
    ])
    df2 = pd.DataFrame([
        {"symbol": "B", "close": 200.0, "price": 200.0, "pct_gain": 10.0},
        {"symbol": "C", "close": 300.0, "price": 300.0, "pct_gain": 15.0},
    ])

    table.populate(df1)
    assert table._populate_in_flight is False
    assert table.model_src.rowCount() == 1

    table.populate(df2)
    assert table._populate_in_flight is False
    assert table.model_src.rowCount() == 2


def test_populate_in_flight_flag_clears_on_inner_exception(_qapp):
    """If `_populate_impl` raises, the flag MUST still clear so the
    next populate isn't permanently dropped. The `try / finally` in
    populate() owns this invariant."""
    import pandas as pd
    from trade_scanner_fh.gui.widgets import ResultsTable

    table = ResultsTable()
    df = pd.DataFrame([
        {"symbol": "A", "close": 100.0, "price": 100.0, "pct_gain": 5.0},
    ])

    def boom(_d):
        raise RuntimeError("simulated render crash")

    table._populate_impl = boom
    try:
        table.populate(df)
    except RuntimeError:
        pass
    # Flag cleared so subsequent populates can proceed
    assert table._populate_in_flight is False


def test_on_timeframe_changed_drops_re_entrant_dropdown_click(_qapp, monkeypatch):
    """End-to-end: simulate the user clicking the dropdown twice in
    quick succession (the second click arrives while the first
    populate is still rendering). The slot's re-entrancy guard must
    drop the inner call rather than recursing — and a follow-up to
    the latest selection must be queued via QTimer.singleShot so the
    most recent click eventually wins WITHOUT synchronous recursion."""
    from unittest.mock import MagicMock, patch
    from trade_scanner_fh.gui.main_window import MainWindow
    import pandas as pd

    parent = MainWindow.__new__(MainWindow)
    # Pre-seed every attribute the slot reads. Bypassing __init__
    # means Qt's attribute-lookup raises if we let `getattr` fall
    # through to the QMainWindow C++ side; setting these explicitly
    # keeps everything in __dict__ where plain Python lookup succeeds.
    parent._timeframe_switch_in_flight = False
    parent._period_results = {
        "1D": pd.DataFrame([{"symbol": "A", "close": 1.0, "price": 1.0, "pct_gain": 1.0}]),
        "1W": pd.DataFrame([{"symbol": "B", "close": 2.0, "price": 2.0, "pct_gain": 2.0}]),
    }
    parent._period_order = ["1D", "1W"]
    parent._results_column_order = []
    parent._active_period = None
    parent._last_results_df = None

    # Simulate the dropdown
    parent.combo_timeframe = MagicMock()
    parent.combo_timeframe.itemData.side_effect = lambda i: ["1D", "1W"][i]
    parent.combo_timeframe.currentIndex.return_value = 0  # for follow-up check

    # Stub the render-side widgets so we can track populate invocations
    populate_calls = []
    parent.results_table = MagicMock()
    parent.results_table.populate.side_effect = lambda df: populate_calls.append(df)
    parent.results_table.set_saved_column_order = MagicMock()

    parent.btn_send = MagicMock()
    parent.btn_export = MagicMock()
    parent.btn_excel = MagicMock()
    parent.status = MagicMock()
    # View-filter checkboxes — default all off so _apply_view_filters
    # is a no-op pass-through. The test exercises re-entrancy, not
    # view-filter behavior.
    parent.chk_view_earnings_dates_only = MagicMock()
    parent.chk_view_earnings_dates_only.isChecked.return_value = False
    parent.chk_view_earnings_data_only = MagicMock()
    parent.chk_view_earnings_data_only.isChecked.return_value = False
    parent.chk_view_color_match_only = MagicMock()
    parent.chk_view_color_match_only.isChecked.return_value = False

    # First click on index 1 — switches to 1W
    parent._on_timeframe_changed(1)
    assert len(populate_calls) == 1
    assert parent._timeframe_switch_in_flight is False

    # Now simulate re-entrancy: set the in-flight flag and try to
    # change again — the second call must be dropped.
    parent._timeframe_switch_in_flight = True
    parent._on_timeframe_changed(0)  # would normally render 1D
    assert len(populate_calls) == 1, "re-entrant call must be dropped"
    parent._timeframe_switch_in_flight = False

    # Sanity: after the flag clears, a normal call works
    parent._on_timeframe_changed(0)
    assert len(populate_calls) == 2


# ──────────────────────────────────────────────────────────────────────
# populate() perf — column-width cache + setUpdatesEnabled wrap
# ──────────────────────────────────────────────────────────────────────

def test_populate_caches_column_widths_after_first_render(_qapp):
    """After the first populate against a given column set, the
    measured widths are snapshotted into `_cached_column_widths`
    keyed by the column-key tuple. Subsequent populates with the
    SAME shape restore widths via setColumnWidth (constant time)
    instead of re-running the per-cell `resizeColumnsToContents`
    measurement loop (which was the dominant cost of the
    timeframe-switch hang the user reported)."""
    import pandas as pd
    from trade_scanner_fh.gui.widgets import ResultsTable

    table = ResultsTable()
    df = pd.DataFrame([
        {"symbol": f"T{i:04d}", "close": 100.0 + i, "price": 100.0 + i,
         "pct_gain": 5.0}
        for i in range(50)
    ])

    assert table._cached_column_widths == {}
    table.populate(df)

    # After first populate: cache has one entry keyed by the column-set
    # tuple, with widths matching the table's actual column widths.
    assert len(table._cached_column_widths) == 1
    cols_key = tuple(c[1] for c in table.active_columns)
    assert cols_key in table._cached_column_widths
    cached = table._cached_column_widths[cols_key]
    assert len(cached) == len(table.active_columns)
    # Widths should be positive integers
    assert all(isinstance(w, int) and w > 0 for w in cached)


def test_populate_uses_cached_widths_on_same_shape_render(_qapp, monkeypatch):
    """Second populate with the same column set must NOT call
    `resizeColumnsToContents` (the slow path). Instead it should
    restore widths from cache via per-column setColumnWidth."""
    import pandas as pd
    from trade_scanner_fh.gui.widgets import ResultsTable

    table = ResultsTable()
    df1 = pd.DataFrame([
        {"symbol": "A", "close": 100.0, "price": 100.0, "pct_gain": 5.0},
    ])
    df2 = pd.DataFrame([
        {"symbol": "B", "close": 200.0, "price": 200.0, "pct_gain": 10.0},
        {"symbol": "C", "close": 300.0, "price": 300.0, "pct_gain": 15.0},
    ])

    # First populate — cold path, measures widths
    table.populate(df1)
    assert len(table._cached_column_widths) == 1

    # Spy on resizeColumnsToContents — should NOT fire on the second
    # populate (same shape, cache hits)
    resize_calls = []
    orig_resize = table.resizeColumnsToContents
    def spy():
        resize_calls.append(True)
        return orig_resize()
    monkeypatch.setattr(table, "resizeColumnsToContents", spy)

    table.populate(df2)
    assert resize_calls == [], (
        "second populate with same column shape must NOT call "
        "resizeColumnsToContents (slow path) — cache should hit"
    )


def test_populate_re_resizes_on_column_set_change(_qapp, monkeypatch):
    """When the column set changes (different filters → different
    columns in df), the cache miss triggers a fresh resize and the
    new shape gets its own cache entry."""
    import pandas as pd
    from trade_scanner_fh.gui.widgets import ResultsTable

    table = ResultsTable()
    df1 = pd.DataFrame([
        {"symbol": "A", "close": 100.0, "price": 100.0, "pct_gain": 5.0},
    ])
    df2 = pd.DataFrame([
        # Add a column that wasn't in df1
        {"symbol": "A", "close": 100.0, "price": 100.0, "pct_gain": 5.0,
         "max_gap_pct": 12.5, "max_gap_date": pd.Timestamp("2025-11-15")},
    ])
    table.populate(df1)
    cache_after_first = dict(table._cached_column_widths)
    assert len(cache_after_first) == 1

    resize_calls = []
    orig_resize = table.resizeColumnsToContents
    monkeypatch.setattr(table, "resizeColumnsToContents",
                        lambda: (resize_calls.append(True), orig_resize())[1])

    table.populate(df2)
    assert len(resize_calls) == 1, "shape changed → resize must fire"
    # Cache now has TWO entries (one per shape)
    assert len(table._cached_column_widths) == 2


def test_populate_setupdatesenabled_wraps_render(_qapp):
    """The setUpdatesEnabled(False)/True wrap is what gives the bulk
    insert its perf — without it, every setItem triggers a partial
    repaint (~38k repaints for a 379×100 render). Verify the wrap
    fires by spying on the toggle calls."""
    import pandas as pd
    from trade_scanner_fh.gui.widgets import ResultsTable

    table = ResultsTable()
    toggle_calls = []
    orig = table.setUpdatesEnabled
    table.setUpdatesEnabled = lambda enabled: (
        toggle_calls.append(enabled), orig(enabled)
    )[1]

    df = pd.DataFrame([
        {"symbol": "A", "close": 100.0, "price": 100.0, "pct_gain": 5.0},
    ])
    table.populate(df)

    # Should have at least one False (start of populate) followed by
    # True (end). Order matters — False first.
    assert False in toggle_calls
    assert True in toggle_calls
    first_false = toggle_calls.index(False)
    last_true = len(toggle_calls) - 1 - toggle_calls[::-1].index(True)
    assert first_false < last_true


def test_populate_detaches_proxy_during_bulk_insert(_qapp):
    """The CRITICAL perf fix: `proxy.setSourceModel(None)` is called
    BEFORE the row loop and `proxy.setSourceModel(self.model_src)` AT
    THE END (in the finally block). Without this detach, the proxy
    receives a `dataChanged` signal for every setItem call (~38k
    signals on a 379×100 render) and processes each one, dominating
    the cost — measured at 124 seconds for the user's actual
    sequenced-run shape.

    With the detach, the bulk insert completes in ~360 ms because
    the proxy gets ONE `modelReset` signal at the end and rebuilds
    its mapping in one pass. This is the difference between an
    instant timeframe switch and a Windows 'Not Responding' hang."""
    import pandas as pd
    from trade_scanner_fh.gui.widgets import ResultsTable

    table = ResultsTable()
    detach_calls = []
    orig = table.proxy.setSourceModel

    def spy(model):
        detach_calls.append(model)
        return orig(model)

    table.proxy.setSourceModel = spy

    df = pd.DataFrame([
        {"symbol": "A", "close": 100.0, "price": 100.0, "pct_gain": 5.0},
        {"symbol": "B", "close": 200.0, "price": 200.0, "pct_gain": 10.0},
    ])
    table.populate(df)

    # Two calls: one with None (detach), one with the source model
    # (reattach). Order matters — detach must come first.
    assert len(detach_calls) >= 2, (
        f"expected at least 2 setSourceModel calls, got {len(detach_calls)}"
    )
    assert None in detach_calls, "proxy.setSourceModel(None) must fire"
    assert table.model_src in detach_calls, (
        "proxy.setSourceModel(model_src) must reattach at end"
    )
    none_idx = detach_calls.index(None)
    src_idx = detach_calls.index(table.model_src)
    assert none_idx < src_idx, (
        "detach must precede reattach (otherwise the proxy stays "
        "attached during the bulk insert and we get the 124-second hang)"
    )

    # And the proxy ends up correctly attached to the source — sanity
    # check that the user's view sees data after populate.
    assert table.proxy.sourceModel() is table.model_src
    assert table.proxy.rowCount() == 2


# ──────────────────────────────────────────────────────────────────────
# View-only filters (post-scan, hide rows in displayed table)
# ──────────────────────────────────────────────────────────────────────

def _shell_with_view_filters(dates_only=False, data_only=False,
                             color_match_only=False):
    """Build a MainWindow shell with the three view-filter checkboxes
    pre-seeded to the requested state. Tests `_apply_view_filters`
    directly without standing up the full GUI."""
    from trade_scanner_fh.gui.main_window import MainWindow
    from unittest.mock import MagicMock
    parent = MainWindow.__new__(MainWindow)
    parent.chk_view_earnings_dates_only = MagicMock()
    parent.chk_view_earnings_dates_only.isChecked.return_value = dates_only
    parent.chk_view_earnings_data_only = MagicMock()
    parent.chk_view_earnings_data_only.isChecked.return_value = data_only
    parent.chk_view_color_match_only = MagicMock()
    parent.chk_view_color_match_only.isChecked.return_value = color_match_only
    return parent


def test_view_filter_no_op_when_all_off(_qapp):
    """With all view filters off, _apply_view_filters returns the
    df unchanged (or a clean copy of it)."""
    import pandas as pd
    parent = _shell_with_view_filters()
    df = pd.DataFrame([
        {"symbol": "A", "close": 1.0, "reported_eps": 2.0},
        {"symbol": "B", "close": 2.0, "reported_eps": float("nan")},
    ])
    out = parent._apply_view_filters(df)
    assert len(out) == 2
    assert set(out["symbol"]) == {"A", "B"}


def test_view_filter_earnings_data_keeps_rows_with_any_earnings_value(_qapp):
    """Earnings Data view filter: a row is kept iff at least one of
    the 6 most-recent-quarter columns OR any q-beats data column is
    non-NaN. Rows with all NaN earnings are dropped."""
    import pandas as pd
    parent = _shell_with_view_filters(data_only=True)

    df = pd.DataFrame([
        # AAPL has reported_eps populated → KEEP
        {"symbol": "AAPL", "close": 100.0, "reported_eps": 2.0,
         "surprise_eps_pct": float("nan"), "reported_rev": float("nan"),
         "surprise_eps_dollar": float("nan"), "surprise_rev_dollar": float("nan"),
         "surprise_rev_pct": float("nan")},
        # MSFT has only revenue surprise → KEEP (any non-NaN counts)
        {"symbol": "MSFT", "close": 200.0, "reported_eps": float("nan"),
         "surprise_eps_pct": float("nan"), "reported_rev": float("nan"),
         "surprise_eps_dollar": float("nan"), "surprise_rev_dollar": 5.0,
         "surprise_rev_pct": float("nan")},
        # NODATA — all 6 NaN → DROP
        {"symbol": "NODATA", "close": 50.0, "reported_eps": float("nan"),
         "surprise_eps_pct": float("nan"), "reported_rev": float("nan"),
         "surprise_eps_dollar": float("nan"), "surprise_rev_dollar": float("nan"),
         "surprise_rev_pct": float("nan")},
    ])
    out = parent._apply_view_filters(df)
    assert set(out["symbol"]) == {"AAPL", "MSFT"}
    assert "NODATA" not in set(out["symbol"])


def test_view_filter_earnings_data_includes_q_beats_columns(_qapp):
    """A ticker with NaN in the 6 most-recent cols but a value in any
    q-beats data column must still pass the Earnings Data filter."""
    import pandas as pd
    parent = _shell_with_view_filters(data_only=True)
    df = pd.DataFrame([
        # Most-recent cols all NaN, but q3_surprise_eps_pct has data
        {"symbol": "BEATER",
         "reported_eps": float("nan"),
         "surprise_eps_dollar": float("nan"),
         "surprise_eps_pct": float("nan"),
         "reported_rev": float("nan"),
         "surprise_rev_dollar": float("nan"),
         "surprise_rev_pct": float("nan"),
         "q3_surprise_eps_pct": 5.0},
        # All earnings cols NaN → DROP
        {"symbol": "EMPTY",
         "reported_eps": float("nan"),
         "surprise_eps_pct": float("nan"),
         "reported_rev": float("nan"),
         "surprise_eps_dollar": float("nan"),
         "surprise_rev_dollar": float("nan"),
         "surprise_rev_pct": float("nan"),
         "q3_surprise_eps_pct": float("nan")},
    ])
    out = parent._apply_view_filters(df)
    assert set(out["symbol"]) == {"BEATER"}


def test_view_filter_earnings_data_when_no_earnings_columns_present(_qapp):
    """If the df has none of the earnings columns at all, the
    Earnings Data view filter eliminates everything (correct: no
    ticker has earnings coverage in this scan)."""
    import pandas as pd
    parent = _shell_with_view_filters(data_only=True)
    df = pd.DataFrame([
        {"symbol": "A", "close": 1.0, "pct_gain": 5.0},
        {"symbol": "B", "close": 2.0, "pct_gain": 10.0},
    ])
    out = parent._apply_view_filters(df)
    assert len(out) == 0


def test_view_filter_earnings_dates_keeps_rows_with_calendar_or_data(_qapp):
    """Earnings Dates view filter: a row is kept iff ANY date column
    (last_report_date / next_earnings_date / days_since_er /
    days_until_er / q*_report_date_*) OR any earnings DATA column is
    populated. The dates filter is broader than the data filter."""
    import pandas as pd
    parent = _shell_with_view_filters(dates_only=True)

    df = pd.DataFrame([
        # Has a calendar date but no data → KEEP (date alone)
        {"symbol": "CAL_ONLY", "last_report_date": pd.Timestamp("2026-02-15"),
         "reported_eps": float("nan")},
        # Has data but no date column populated → KEEP (data implies date)
        {"symbol": "DATA_ONLY", "last_report_date": pd.NaT,
         "reported_eps": 1.5},
        # Has both → KEEP
        {"symbol": "BOTH", "last_report_date": pd.Timestamp("2026-02-15"),
         "reported_eps": 2.0},
        # Has neither → DROP
        {"symbol": "NEITHER", "last_report_date": pd.NaT,
         "reported_eps": float("nan")},
    ])
    out = parent._apply_view_filters(df)
    assert set(out["symbol"]) == {"CAL_ONLY", "DATA_ONLY", "BOTH"}


def test_view_filter_dates_supersets_data(_qapp):
    """Invariant check: any row passing the Data filter also passes
    the Dates filter (data implies date). Run both filters separately
    on the same df and verify dates output ⊇ data output."""
    import pandas as pd
    df = pd.DataFrame([
        {"symbol": "A", "last_report_date": pd.Timestamp("2026-02-15"),
         "reported_eps": float("nan")},
        {"symbol": "B", "last_report_date": pd.NaT,
         "reported_eps": 1.5},
        {"symbol": "C", "last_report_date": pd.Timestamp("2026-02-15"),
         "reported_eps": 1.5},
        {"symbol": "D", "last_report_date": pd.NaT,
         "reported_eps": float("nan")},
    ])

    parent_dates = _shell_with_view_filters(dates_only=True)
    parent_data = _shell_with_view_filters(data_only=True)
    out_dates = set(parent_dates._apply_view_filters(df)["symbol"])
    out_data = set(parent_data._apply_view_filters(df)["symbol"])
    assert out_data.issubset(out_dates), (
        f"data filter ({out_data}) must be a subset of dates filter ({out_dates})"
    )


def test_view_filter_color_match_only_keeps_rows_with_aligned_dates(_qapp):
    """Color Match Only view filter: a row is kept iff
    `_earnings_aligned_dates` is a non-empty list (the scanner's
    'this ticker had a non-earnings indicator land on an earnings
    date' signal)."""
    import pandas as pd
    parent = _shell_with_view_filters(color_match_only=True)
    df = pd.DataFrame([
        {"symbol": "MATCH",   "close": 1.0,
         "_earnings_aligned_dates": ["2026-02-15"]},
        {"symbol": "EMPTY",   "close": 2.0,
         "_earnings_aligned_dates": []},
        {"symbol": "MISSING", "close": 3.0,
         "_earnings_aligned_dates": float("nan")},
        {"symbol": "ANOTHER", "close": 4.0,
         "_earnings_aligned_dates": ["2025-08-15", "2025-11-15"]},
    ])
    out = parent._apply_view_filters(df)
    assert set(out["symbol"]) == {"MATCH", "ANOTHER"}


def test_view_filter_color_match_only_when_column_missing(_qapp):
    """If the `_earnings_aligned_dates` column itself doesn't exist
    in the df, the Color Match Only filter eliminates everything."""
    import pandas as pd
    parent = _shell_with_view_filters(color_match_only=True)
    df = pd.DataFrame([
        {"symbol": "A", "close": 1.0},
        {"symbol": "B", "close": 2.0},
    ])
    out = parent._apply_view_filters(df)
    assert len(out) == 0


def test_view_filter_data_and_color_match_intersect(_qapp):
    """When Earnings Data + Color Match Only are both on, only rows
    passing BOTH survive."""
    import pandas as pd
    parent = _shell_with_view_filters(data_only=True, color_match_only=True)
    df = pd.DataFrame([
        {"symbol": "EARN_NO_MATCH", "reported_eps": 2.0,
         "_earnings_aligned_dates": []},
        {"symbol": "MATCH_NO_EARN", "reported_eps": float("nan"),
         "_earnings_aligned_dates": ["2026-02-15"]},
        {"symbol": "BOTH", "reported_eps": 2.0,
         "_earnings_aligned_dates": ["2026-02-15"]},
        {"symbol": "NEITHER", "reported_eps": float("nan"),
         "_earnings_aligned_dates": []},
    ])
    out = parent._apply_view_filters(df)
    assert set(out["symbol"]) == {"BOTH"}


def test_view_filter_returns_empty_df_unchanged(_qapp):
    """An empty / None df is returned without crashing. Important
    because populate is called with empty frames during state
    transitions (initial GUI startup, post-scan-with-zero-results)."""
    import pandas as pd
    parent = _shell_with_view_filters(data_only=True)
    out_empty = parent._apply_view_filters(pd.DataFrame())
    assert out_empty is not None and len(out_empty) == 0
    out_none = parent._apply_view_filters(None)
    assert out_none is None


def test_view_filter_returned_df_has_clean_index(_qapp):
    """After dropping rows, the returned df must have a clean
    RangeIndex so downstream code (Excel writer, get_symbols) doesn't
    trip on missing index entries."""
    import pandas as pd
    parent = _shell_with_view_filters(data_only=True)
    df = pd.DataFrame([
        {"symbol": "DROP",  "reported_eps": float("nan")},  # dropped
        {"symbol": "KEEP",  "reported_eps": 1.5},            # kept
        {"symbol": "DROP2", "reported_eps": float("nan")},  # dropped
        {"symbol": "KEEP2", "reported_eps": 2.5},            # kept
    ])
    out = parent._apply_view_filters(df)
    assert list(out.index) == [0, 1]
    assert list(out["symbol"]) == ["KEEP", "KEEP2"]


def test_populate_proxy_reattaches_even_on_inner_exception(_qapp):
    """If `_populate_impl` raises mid-loop, the proxy reattach in
    the `finally` block MUST still fire — otherwise the table view
    is permanently stuck pointing at None and shows blank forever."""
    import pandas as pd
    from trade_scanner_fh.gui.widgets import ResultsTable

    table = ResultsTable()
    df = pd.DataFrame([
        {"symbol": "A", "close": 100.0, "price": 100.0, "pct_gain": 5.0},
    ])

    # First populate to ensure proxy is attached
    table.populate(df)
    assert table.proxy.sourceModel() is table.model_src

    # Sabotage _populate_row so it raises mid-loop
    def boom(*_args, **_kw):
        raise RuntimeError("simulated mid-loop crash")
    table._populate_row = boom

    # populate() catches per-row exceptions inside the loop, so this
    # actually completes (no exception escapes). But proxy must still
    # be reattached at the end.
    table.populate(df)
    assert table.proxy.sourceModel() is table.model_src, (
        "proxy reattach must happen even when row rendering fails"
    )


# ──────────────────────────────────────────────────────────────────────
# Delete rows from output window — context menu + Delete key handler
# ──────────────────────────────────────────────────────────────────────

def test_results_table_delete_key_emits_deletion_request(_qapp):
    """Delete key on a selected row emits rows_deletion_requested with
    the row's symbol. Multi-select returns multiple symbols in
    selection order."""
    from PyQt6.QtCore import Qt as _Qt
    from PyQt6.QtGui import QKeyEvent
    from PyQt6.QtCore import QEvent
    from trade_scanner_fh.gui.widgets import ResultsTable

    df = pd.DataFrame([
        {"symbol": "AAPL", "close": 1.0, "pct_gain": 1.0},
        {"symbol": "MSFT", "close": 2.0, "pct_gain": 2.0},
        {"symbol": "NVDA", "close": 3.0, "pct_gain": 3.0},
    ])
    table = ResultsTable()
    table.populate(df)

    captured: list[list[str]] = []
    table.rows_deletion_requested.connect(captured.append)

    # Select MSFT (proxy-row 1) — assume identity sort.
    sel = table.selectionModel()
    proxy_row_1 = table.proxy.index(1, 0)
    sel.select(
        proxy_row_1,
        sel.SelectionFlag.Select | sel.SelectionFlag.Rows,
    )

    evt = QKeyEvent(QEvent.Type.KeyPress, _Qt.Key.Key_Delete, _Qt.KeyboardModifier.NoModifier)
    table.keyPressEvent(evt)

    assert captured == [["MSFT"]]


def test_results_table_delete_multiselect_returns_all_symbols(_qapp):
    from PyQt6.QtCore import Qt as _Qt
    from PyQt6.QtGui import QKeyEvent
    from PyQt6.QtCore import QEvent
    from trade_scanner_fh.gui.widgets import ResultsTable

    df = pd.DataFrame([
        {"symbol": "AAPL", "close": 1.0, "pct_gain": 1.0},
        {"symbol": "MSFT", "close": 2.0, "pct_gain": 2.0},
        {"symbol": "NVDA", "close": 3.0, "pct_gain": 3.0},
    ])
    table = ResultsTable()
    table.populate(df)

    captured: list[list[str]] = []
    table.rows_deletion_requested.connect(captured.append)

    sel = table.selectionModel()
    for r in (0, 2):
        sel.select(
            table.proxy.index(r, 0),
            sel.SelectionFlag.Select | sel.SelectionFlag.Rows,
        )

    evt = QKeyEvent(QEvent.Type.KeyPress, _Qt.Key.Key_Delete, _Qt.KeyboardModifier.NoModifier)
    table.keyPressEvent(evt)

    assert len(captured) == 1
    assert set(captured[0]) == {"AAPL", "NVDA"}


def test_results_table_delete_with_no_selection_is_noop(_qapp):
    """Delete key with no selected row must not emit (no symbols to
    delete) — falls through to default handler."""
    from PyQt6.QtCore import Qt as _Qt
    from PyQt6.QtGui import QKeyEvent
    from PyQt6.QtCore import QEvent
    from trade_scanner_fh.gui.widgets import ResultsTable

    df = pd.DataFrame([{"symbol": "A", "close": 1.0, "pct_gain": 1.0}])
    table = ResultsTable()
    table.populate(df)

    captured: list[list[str]] = []
    table.rows_deletion_requested.connect(captured.append)

    table.selectionModel().clearSelection()
    evt = QKeyEvent(QEvent.Type.KeyPress, _Qt.Key.Key_Delete, _Qt.KeyboardModifier.NoModifier)
    table.keyPressEvent(evt)

    assert captured == []


# ──────────────────────────────────────────────────────────────────────
# Excel export: include-cell-colors checkbox + color application
# ──────────────────────────────────────────────────────────────────────

def test_export_dialog_colors_checkbox_grey_for_csv(_qapp):
    """The 'Include cell colors' checkbox is enabled for XLSX and
    disabled (greyed out) when CSV is the selected format."""
    from trade_scanner_fh.gui.dialogs import ExcelExportDialog

    dlg = ExcelExportDialog(
        columns=[("Symbol", "symbol", str)],
        periods=["P1"],
    )
    # XLSX is the default — checkbox should be enabled
    assert dlg._chk_colors.isEnabled()

    dlg._combo_format.setCurrentText("CSV")
    assert not dlg._chk_colors.isEnabled()
    # wants_colors returns False for CSV regardless of check state
    assert dlg.wants_colors() is False

    dlg._combo_format.setCurrentText("XLSX")
    assert dlg._chk_colors.isEnabled()


def test_export_dialog_wants_colors_default_true_xlsx(_qapp):
    from trade_scanner_fh.gui.dialogs import ExcelExportDialog
    dlg = ExcelExportDialog(
        columns=[("Symbol", "symbol", str)],
        periods=["P1"],
    )
    assert dlg.format_choice() == "xlsx"
    assert dlg.wants_colors() is True


# ──────────────────────────────────────────────────────────────────────
# Interleave Quarters layout — view-only column rearrangement when
# both EPS and Rev consec-beats are active. No-op when only one side
# is present so back-compat is preserved.
# ──────────────────────────────────────────────────────────────────────

def test_interleave_quarters_groups_eps_and_rev_per_quarter(_qapp):
    """When both EPS and Rev beats are populated AND the toggle is on,
    Q-i blocks alternate (Q-1 EPS+Rev, then Q-2 EPS+Rev, ...) so all
    data for one quarter sits adjacent."""
    from trade_scanner_fh.gui.widgets import _build_dynamic_columns

    df = pd.DataFrame([{
        "symbol": "A", "close": 1.0, "pct_gain": 1.0,
        "consec_eps_beats": 2, "consec_rev_beats": 2,
        "q1_report_date_eps": pd.Timestamp("2026-02-15"),
        "q1_reported_eps": 2.0, "q1_surprise_eps_dollar": 0.1,
        "q1_surprise_eps_pct": 5.0, "q1_yoy_eps_pct": 10.0,
        "q2_report_date_eps": pd.Timestamp("2025-11-15"),
        "q2_reported_eps": 1.9, "q2_surprise_eps_dollar": 0.05,
        "q2_surprise_eps_pct": 3.0, "q2_yoy_eps_pct": 8.0,
        "q1_report_date_rev": pd.Timestamp("2026-02-15"),
        "q1_reported_rev": 100.0, "q1_surprise_rev_dollar": 5.0,
        "q1_surprise_rev_pct": 5.0, "q1_yoy_rev_pct": 12.0,
        "q2_report_date_rev": pd.Timestamp("2025-11-15"),
        "q2_reported_rev": 95.0, "q2_surprise_rev_dollar": 3.0,
        "q2_surprise_rev_pct": 3.0, "q2_yoy_rev_pct": 9.0,
    }])
    cols, _, _ = _build_dynamic_columns(df, interleave_quarters=True)
    keys = [k for _h, k, _f in cols]
    # Both Consec counters up front (adjacent)
    eps_counter = keys.index("consec_eps_beats")
    rev_counter = keys.index("consec_rev_beats")
    assert rev_counter == eps_counter + 1
    # Q-1 EPS block precedes Q-1 Rev block precedes Q-2 EPS block
    q1_eps = keys.index("q1_report_date_eps")
    q1_rev = keys.index("q1_report_date_rev")
    q2_eps = keys.index("q2_report_date_eps")
    q2_rev = keys.index("q2_report_date_rev")
    assert q1_eps < q1_rev < q2_eps < q2_rev


def test_interleave_quarters_no_op_when_only_eps(_qapp):
    """Single-side beats: interleave flag must NOT change the layout
    (back-compat — guarantees zero behavior change for users who run
    EPS-only or Rev-only beats scans)."""
    from trade_scanner_fh.gui.widgets import _build_dynamic_columns

    df = pd.DataFrame([{
        "symbol": "A", "close": 1.0, "pct_gain": 1.0,
        "consec_eps_beats": 1,
        "q1_report_date_eps": pd.Timestamp("2026-02-15"),
        "q1_reported_eps": 2.0, "q1_surprise_eps_dollar": 0.1,
        "q1_surprise_eps_pct": 5.0,
    }])
    cols_off, _, _ = _build_dynamic_columns(df, interleave_quarters=False)
    cols_on, _, _ = _build_dynamic_columns(df, interleave_quarters=True)
    assert [k for _h, k, _f in cols_off] == [k for _h, k, _f in cols_on]


def test_interleave_quarters_no_op_when_only_rev(_qapp):
    from trade_scanner_fh.gui.widgets import _build_dynamic_columns

    df = pd.DataFrame([{
        "symbol": "A", "close": 1.0, "pct_gain": 1.0,
        "consec_rev_beats": 1,
        "q1_report_date_rev": pd.Timestamp("2026-02-15"),
        "q1_reported_rev": 100.0, "q1_surprise_rev_dollar": 5.0,
        "q1_surprise_rev_pct": 5.0,
    }])
    cols_off, _, _ = _build_dynamic_columns(df, interleave_quarters=False)
    cols_on, _, _ = _build_dynamic_columns(df, interleave_quarters=True)
    assert [k for _h, k, _f in cols_off] == [k for _h, k, _f in cols_on]


def test_interleave_quarters_handles_asymmetric_n_eps_n_rev(_qapp):
    """n_eps != n_rev — interleave emits whichever side still has data
    at each quarter index. The shorter side ends; the longer side
    continues with its remaining quarters."""
    from trade_scanner_fh.gui.widgets import _build_dynamic_columns

    df = pd.DataFrame([{
        "symbol": "A", "close": 1.0, "pct_gain": 1.0,
        "consec_eps_beats": 1, "consec_rev_beats": 3,
        "q1_report_date_eps": pd.Timestamp("2026-02-15"),
        "q1_reported_eps": 2.0, "q1_surprise_eps_dollar": 0.1,
        "q1_surprise_eps_pct": 5.0,
        "q1_report_date_rev": pd.Timestamp("2026-02-15"),
        "q1_reported_rev": 100.0, "q1_surprise_rev_dollar": 5.0,
        "q1_surprise_rev_pct": 5.0,
        "q2_report_date_rev": pd.Timestamp("2025-11-15"),
        "q2_reported_rev": 95.0, "q2_surprise_rev_dollar": 3.0,
        "q2_surprise_rev_pct": 3.0,
        "q3_report_date_rev": pd.Timestamp("2025-08-15"),
        "q3_reported_rev": 90.0, "q3_surprise_rev_dollar": 2.0,
        "q3_surprise_rev_pct": 2.0,
    }])
    cols, n_eps, n_rev = _build_dynamic_columns(df, interleave_quarters=True)
    keys = [k for _h, k, _f in cols]
    assert n_eps == 1 and n_rev == 3
    assert keys.index("q1_report_date_eps") < keys.index("q1_report_date_rev")
    assert keys.index("q1_report_date_rev") < keys.index("q2_report_date_rev")
    assert keys.index("q2_report_date_rev") < keys.index("q3_report_date_rev")
    # No q2 EPS columns should be present
    assert "q2_report_date_eps" not in keys


def test_results_table_set_interleave_quarters_setter(_qapp):
    """ResultsTable exposes set_interleave_quarters(bool); it flips the
    flag and is no-op on no-change."""
    from trade_scanner_fh.gui.widgets import ResultsTable
    table = ResultsTable()
    assert table.interleave_quarters is False
    table.set_interleave_quarters(True)
    assert table.interleave_quarters is True
    table.set_interleave_quarters(False)
    assert table.interleave_quarters is False


# ──────────────────────────────────────────────────────────────────────
# Cut/Paste rows — clipboard + reorder via MainWindow paste handler.
# ──────────────────────────────────────────────────────────────────────

def test_cut_clipboard_set_and_clear(_qapp):
    from trade_scanner_fh.gui.widgets import ResultsTable
    table = ResultsTable()
    df = pd.DataFrame([{"symbol": "AAPL", "close": 1.0, "pct_gain": 1.0}])
    table.populate(df)
    assert table.cut_clipboard() == []
    table._set_cut_clipboard(["AAPL", "MSFT"])
    assert table.cut_clipboard() == ["AAPL", "MSFT"]
    table.clear_cut_clipboard()
    assert table.cut_clipboard() == []


def test_fire_paste_rejects_target_in_cut_set(_qapp):
    """Pasting onto a row that's part of the cut set is silently
    ignored (would orphan the target). No signal emitted."""
    from trade_scanner_fh.gui.widgets import ResultsTable
    table = ResultsTable()
    captured = []
    table.rows_paste_requested.connect(lambda c, t: captured.append((c, t)))
    table._fire_paste(["AAPL", "MSFT"], "AAPL")
    assert captured == []


def test_fire_paste_emits_signal_on_valid_target(_qapp):
    from trade_scanner_fh.gui.widgets import ResultsTable
    table = ResultsTable()
    captured = []
    table.rows_paste_requested.connect(lambda c, t: captured.append((c, t)))
    table._fire_paste(["AAPL", "MSFT"], "NVDA")
    assert captured == [(["AAPL", "MSFT"], "NVDA")]


def _shell_for_paste_tests(_qapp, df):
    """Build a MainWindow shell with just enough state for the paste/
    delete row tests to drive `_on_rows_paste_requested` /
    `_on_rows_deletion_requested` directly. Bypasses the full GUI
    init (which leaks log handlers across tests)."""
    from trade_scanner_fh.gui.main_window import MainWindow
    from trade_scanner_fh.gui.widgets import ResultsTable
    parent = MainWindow.__new__(MainWindow)
    parent._period_results = {"P1": df.copy()}
    parent._period_order = ["P1"]
    parent._active_period = "P1"
    parent.results_table = ResultsTable()
    parent.results_table.populate(df)
    return parent


def test_main_window_paste_reorders_rows_after_target(_qapp):
    """MainWindow's paste slot moves the cut block to immediately
    after the target row in `_period_results[active_period]`. Order
    of cut symbols is preserved as the user originally cut them."""
    df = pd.DataFrame([
        {"symbol": s, "close": 1.0, "pct_gain": 1.0}
        for s in ["AAPL", "MSFT", "NVDA", "GOOG", "TSLA"]
    ])
    win = _shell_for_paste_tests(_qapp, df)
    win._on_rows_paste_requested(["AAPL", "NVDA"], "TSLA")
    new_order = list(win._period_results["P1"]["symbol"])
    # Original: AAPL, MSFT, NVDA, GOOG, TSLA
    # After cut AAPL+NVDA, remaining: MSFT, GOOG, TSLA
    # After paste AAPL+NVDA after TSLA: MSFT, GOOG, TSLA, AAPL, NVDA
    assert new_order == ["MSFT", "GOOG", "TSLA", "AAPL", "NVDA"]


def test_main_window_paste_in_middle_of_remaining_rows(_qapp):
    df = pd.DataFrame([
        {"symbol": s, "close": 1.0, "pct_gain": 1.0}
        for s in ["AAPL", "MSFT", "NVDA", "GOOG", "TSLA"]
    ])
    win = _shell_for_paste_tests(_qapp, df)
    win._on_rows_paste_requested(["NVDA"], "AAPL")
    new_order = list(win._period_results["P1"]["symbol"])
    # Cut NVDA → AAPL, MSFT, GOOG, TSLA
    # Paste NVDA after AAPL → AAPL, NVDA, MSFT, GOOG, TSLA
    assert new_order == ["AAPL", "NVDA", "MSFT", "GOOG", "TSLA"]


def test_main_window_paste_with_unknown_target_is_noop(_qapp):
    """Target row not in df → paste no-op + clipboard cleared."""
    df = pd.DataFrame([
        {"symbol": s, "close": 1.0, "pct_gain": 1.0}
        for s in ["AAPL", "MSFT", "NVDA"]
    ])
    win = _shell_for_paste_tests(_qapp, df)
    win.results_table._set_cut_clipboard(["AAPL"])
    win._on_rows_paste_requested(["AAPL"], "GHOST")
    new_order = list(win._period_results["P1"]["symbol"])
    assert new_order == ["AAPL", "MSFT", "NVDA"]
    assert win.results_table.cut_clipboard() == []


def test_delete_clears_cut_clipboard(_qapp):
    """Row deletion clears the clipboard so paste can't target a
    deleted row."""
    df = pd.DataFrame([
        {"symbol": s, "close": 1.0, "pct_gain": 1.0}
        for s in ["AAPL", "MSFT", "NVDA"]
    ])
    win = _shell_for_paste_tests(_qapp, df)
    win.results_table._set_cut_clipboard(["NVDA"])
    win._on_rows_deletion_requested(["MSFT"])
    assert win.results_table.cut_clipboard() == []


# ──────────────────────────────────────────────────────────────────────
# Delete columns — header right-click → drop from rendered df without
# touching `_period_results`.
# ──────────────────────────────────────────────────────────────────────

def test_header_translator_filters_always_visible(_qapp):
    """The ResultsTable's header→key translator must filter out the
    always-visible core columns (symbol/close/pct_gain/gain_start_date)
    even if the header reports their logical indices — deleting them
    would break export and core display invariants."""
    from trade_scanner_fh.gui.widgets import ResultsTable

    df = pd.DataFrame([{
        "symbol": "A", "close": 100.0, "pct_gain": 5.0,
        "avg_vol": 1000000.0, "sti": 1.5,
    }])
    table = ResultsTable()
    table.populate(df)
    captured = []
    table.columns_deletion_requested.connect(captured.append)

    keys_in_table = [k for _h, k, _f in table.active_columns]
    sym_idx = keys_in_table.index("symbol")
    avg_vol_idx = keys_in_table.index("avg_vol")
    sti_idx = keys_in_table.index("sti")
    table._on_header_columns_deletion_requested(
        [sym_idx, avg_vol_idx, sti_idx]
    )
    assert captured == [["avg_vol", "sti"]]


def _shell_with_deleted_columns(_qapp, deleted_keys: set):
    """Shell for column-delete tests — reuses the existing view-filter
    shell pattern + seeds `_deleted_column_keys`."""
    parent = _shell_with_view_filters()
    parent._deleted_column_keys = set(deleted_keys)
    return parent


def test_main_window_apply_view_filters_drops_deleted_columns(_qapp):
    parent = _shell_with_deleted_columns(_qapp, {"avg_vol", "sti"})
    df = pd.DataFrame([{
        "symbol": "A", "close": 100.0, "pct_gain": 5.0,
        "avg_vol": 1000000.0, "sti": 1.5, "dollar_vol": 99.0,
    }])
    out = parent._apply_view_filters(df)
    assert "symbol" in out.columns
    assert "avg_vol" not in out.columns
    assert "sti" not in out.columns
    assert "dollar_vol" in out.columns


def test_apply_view_filters_tolerates_deleted_col_not_in_df(_qapp):
    """User hid a column that doesn't exist in the current frame
    (e.g. the underlying filter is now off). Should not raise."""
    parent = _shell_with_deleted_columns(_qapp, {"avg_vol", "sti"})
    df = pd.DataFrame([{
        "symbol": "A", "close": 100.0, "pct_gain": 5.0,
        "dollar_vol": 99.0,
    }])
    out = parent._apply_view_filters(df)
    assert "dollar_vol" in out.columns


def test_excel_dialog_bundles_all_qi_eps_suffixes(_qapp):
    """Regression: q-i columns with `_yoy_eps_pct` and `_report_date_eps`
    suffixes were leaking into individual checkboxes instead of the
    EPS bundle group. After fix, ALL q-i EPS variants land in the
    bundle so the user gets a single toggle for the whole block."""
    from trade_scanner_fh.gui.dialogs import ExcelExportDialog

    cols = [
        ("Ticker", "symbol", str),
        ("Q-1 Date", "q1_report_date_eps", str),
        ("Q-1 Reported EPS", "q1_reported_eps", str),
        ("Q-1 Surp EPS $", "q1_surprise_eps_dollar", str),
        ("Q-1 Surp EPS %", "q1_surprise_eps_pct", str),
        ("Q-1 YoY EPS %", "q1_yoy_eps_pct", str),
        ("Q-1 Date (rev)", "q1_report_date_rev", str),
        ("Q-1 Reported Rev", "q1_reported_rev", str),
        ("Q-1 Surp Rev $", "q1_surprise_rev_dollar", str),
        ("Q-1 Surp Rev %", "q1_surprise_rev_pct", str),
        ("Q-1 YoY Rev %", "q1_yoy_rev_pct", str),
    ]
    dlg = ExcelExportDialog(cols, periods=["P1"])

    assert set(dlg._group_eps_keys) == {
        "q1_report_date_eps", "q1_reported_eps",
        "q1_surprise_eps_dollar", "q1_surprise_eps_pct",
        "q1_yoy_eps_pct",
    }
    assert set(dlg._group_rev_keys) == {
        "q1_report_date_rev", "q1_reported_rev",
        "q1_surprise_rev_dollar", "q1_surprise_rev_pct",
        "q1_yoy_rev_pct",
    }
    # No q-i column should leak into regular checkboxes.
    leaked = [k for k in dlg._checks if k.startswith("q")]
    assert leaked == [], (
        f"q-i columns leaked into regular checkboxes: {leaked}"
    )


def test_build_export_df_includes_qi_columns_when_bundle_checked(_qapp):
    """Regression: per-quarter beats columns were missing from XLSX/
    CSV exports because `_build_export_df` only iterated the static
    `RESULT_COLUMNS`. After fix, the function uses the live active
    column layout (which includes dynamic q-i columns) so q-i values
    flow through to the export when their bundle is checked."""
    from trade_scanner_fh.gui.main_window import MainWindow
    from trade_scanner_fh.gui.widgets import ResultsTable

    df = pd.DataFrame([{
        "symbol": "A", "close": 1.0, "pct_gain": 1.0,
        "consec_eps_beats": 1,
        "q1_report_date_eps": pd.Timestamp("2026-02-15"),
        "q1_reported_eps": 2.0,
        "q1_surprise_eps_dollar": 0.1,
        "q1_surprise_eps_pct": 5.0,
        "q1_yoy_eps_pct": 21.0,
    }])
    parent = MainWindow.__new__(MainWindow)
    parent.results_table = ResultsTable()
    parent.results_table.populate(df)

    keys = [
        "symbol", "close", "pct_gain",
        "q1_report_date_eps", "q1_reported_eps",
        "q1_surprise_eps_dollar", "q1_surprise_eps_pct",
        "q1_yoy_eps_pct",
    ]
    out = MainWindow._build_export_df(parent, df, keys, wants_news=False)
    # All q-i columns must surface in the export
    headers = list(out.columns)
    assert "Q-1 Reported EPS" in headers
    assert "Q-1 Surp EPS $" in headers
    assert "Q-1 Surp EPS %" in headers
    assert "Q-1 YoY EPS %" in headers
    assert "Q-1 Date" in headers


def test_interleave_toggle_overrides_saved_column_order(_qapp):
    """Regression: after the user has dragged any column (which
    populates the saved column order), toggling Interleave Q EPS+Rev
    must still rearrange the per-quarter blocks. Previously the saved
    order's `_apply_saved_order` ran AFTER the new interleaved
    `_active_columns` were assigned, undoing the layout flip and
    leaving columns in the prior (non-interleaved) visual order."""
    from trade_scanner_fh.gui.main_window import MainWindow

    df = pd.DataFrame([{
        "symbol": "A", "close": 1.0, "pct_gain": 1.0,
        "consec_eps_beats": 2, "consec_rev_beats": 2,
        "q1_report_date_eps": pd.Timestamp("2026-02-15"),
        "q1_reported_eps": 2.0, "q1_surprise_eps_dollar": 0.1,
        "q1_surprise_eps_pct": 5.0, "q1_yoy_eps_pct": 10.0,
        "q2_report_date_eps": pd.Timestamp("2025-11-15"),
        "q2_reported_eps": 1.9, "q2_surprise_eps_dollar": 0.05,
        "q2_surprise_eps_pct": 3.0, "q2_yoy_eps_pct": 8.0,
        "q1_report_date_rev": pd.Timestamp("2026-02-15"),
        "q1_reported_rev": 100.0, "q1_surprise_rev_dollar": 5.0,
        "q1_surprise_rev_pct": 5.0, "q1_yoy_rev_pct": 12.0,
        "q2_report_date_rev": pd.Timestamp("2025-11-15"),
        "q2_reported_rev": 95.0, "q2_surprise_rev_dollar": 3.0,
        "q2_surprise_rev_pct": 3.0, "q2_yoy_rev_pct": 9.0,
    }])

    mw = MainWindow()
    try:
        mw._period_order = ["live"]
        mw._period_results = {"live": df}
        mw._active_period = "live"
        mw.combo_timeframe.clear()
        mw.combo_timeframe.addItem("live", "live")
        mw.combo_timeframe.setCurrentIndex(0)
        mw._on_timeframe_changed(0)

        # Simulate a column drag — Qt's sectionMoved → order_changed →
        # MainWindow's _results_column_order. Fakes the user having
        # touched the layout before reaching for the interleave toggle.
        header = mw.results_table.horizontalHeader()
        header.moveSection(0, 5)
        assert mw._results_column_order, (
            "drag should have populated _results_column_order"
        )

        # Now toggle interleave on. Setter route fires the slot.
        mw.chk_view_interleave_quarters.setChecked(True)

        # Walk the visual order; q-i blocks must alternate
        # EPS / Rev per quarter rather than EPS-then-all-Rev.
        keys_visual: list[str] = []
        for v in range(header.count()):
            li = header.logicalIndex(v)
            if li < len(mw.results_table._active_columns):
                keys_visual.append(mw.results_table._active_columns[li][1])
        q1_eps = keys_visual.index("q1_report_date_eps")
        q1_rev = keys_visual.index("q1_report_date_rev")
        q2_eps = keys_visual.index("q2_report_date_eps")
        q2_rev = keys_visual.index("q2_report_date_rev")
        assert q1_eps < q1_rev < q2_eps < q2_rev, (
            f"interleave layout not applied; visual q-cols: "
            f"{[k for k in keys_visual if k.startswith('q')]}"
        )
    finally:
        mw.close()
        mw.deleteLater()


def test_header_column_cut_paste_reorders_block_after_target(_qapp):
    """Header right-click → Cut N columns + Paste after target moves
    the cut block to immediately after the target column. Verified
    end-to-end via the helper that the menu wires up: cut clipboard
    + visual-index compensation + `_move_block_to_visual`."""
    from trade_scanner_fh.gui.widgets import ResultsTable
    df = pd.DataFrame([{
        "symbol": "A", "close": 1.0, "pct_gain": 1.0,
        "avg_vol": 100, "sti": 1.5, "dollar_vol": 99, "rs_market": 1.2,
    }])
    table = ResultsTable()
    table.populate(df)
    hdr = table._reorderable_header
    keys = [k for _h, k, _f in table.active_columns]

    def visual_order():
        return [keys[hdr.logicalIndex(v)] for v in range(hdr.count())]

    avg_vol_li = keys.index("avg_vol")
    sti_li = keys.index("sti")
    dollar_vol_li = keys.index("dollar_vol")

    # Cut [sti, avg_vol] (logical), paste after dollar_vol
    cut = [avg_vol_li, sti_li]
    target_visual_now = hdr.visualIndex(dollar_vol_li)
    cut_to_left = sum(
        1 for li in cut if hdr.visualIndex(li) < target_visual_now
    )
    target_visual = target_visual_now - cut_to_left + 1
    hdr._move_block_to_visual(cut, target_visual)

    order = visual_order()
    dvi = order.index("dollar_vol")
    # Block lands immediately after dollar_vol
    assert order[dvi + 1] in ("avg_vol", "sti")
    assert order[dvi + 2] in ("avg_vol", "sti")
    # rs_market trails the block
    assert order[dvi + 3] == "rs_market"


def test_apply_view_filters_no_deleted_columns_passthrough(_qapp):
    """Empty `_deleted_column_keys` — view filter is a no-op on the
    column dimension. Sanity check that the new code path doesn't
    drop anything when the set is empty."""
    parent = _shell_with_deleted_columns(_qapp, set())
    df = pd.DataFrame([{
        "symbol": "A", "close": 100.0, "pct_gain": 5.0,
        "avg_vol": 1000000.0, "sti": 1.5,
    }])
    out = parent._apply_view_filters(df)
    assert set(out.columns) == {"symbol", "close", "pct_gain", "avg_vol", "sti"}


def test_export_xlsx_with_colors_writes_palette_color_to_cell(tmp_path, _qapp):
    """Smoke: write an xlsx with apply_colors=True, then read it back
    via openpyxl and verify a cell that was palette-colored on screen
    got a non-default font color in the workbook."""
    from openpyxl import load_workbook
    from trade_scanner_fh.gui.main_window import MainWindow
    from trade_scanner_fh.gui.widgets import ResultsTable, RESULT_COLUMNS

    same = pd.Timestamp("2026-02-15")
    df = pd.DataFrame([{
        "symbol": "TKR", "close": 100.0, "pct_gain": 5.0,
        "max_gap_pct": 12.5,
        "max_gap_date": same,
        "last_report_date": same,
        "_earnings_aligned_dates": ["2026-02-15"],
    }])

    # Construct a minimal MainWindow shell wired enough for the export
    # path. We bypass the full init by populating the table directly
    # and stubbing the period state.
    win = MainWindow()
    try:
        win._period_results = {"P1": df}
        win._period_order = ["P1"]
        win._active_period = "P1"
        win.results_table.populate(df)
        path = str(tmp_path / "out.xlsx")
        keys = [k for _h, k, _f in RESULT_COLUMNS if k in df.columns]
        win._write_xlsx_multi_sheet(
            path, ["P1"], keys, wants_news=False, apply_colors=True,
        )
        wb = load_workbook(path)
        ws = wb.active
        # Find the column whose header is "Max Gap Date".
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_idx = header_row.index("Max Gap Date") + 1
        # Row 2 is the data row.
        cell = ws.cell(row=2, column=col_idx)
        # Must have a font color set (palette match), not the default.
        assert cell.font.color is not None
        # Palette colors are stored as ARGB hex (FF prefix).
        assert cell.font.color.rgb is not None
    finally:
        win.close()
        win.deleteLater()


# ──────────────────────────────────────────────────────────────────────
# Excel export honors user's manual column drag order
# ──────────────────────────────────────────────────────────────────────

def test_excel_export_honors_manual_column_drag_order(_qapp):
    """Regression: dragging a column in the table must propagate
    through the Excel export. Previously `_build_export_df` iterated
    `active_columns` (canonical order) and only checked membership in
    `keys`, ignoring the user's visual order entirely."""
    from trade_scanner_fh.gui.main_window import MainWindow
    from trade_scanner_fh.gui.widgets import ResultsTable

    df = pd.DataFrame([{
        "symbol": "A", "close": 100.0, "pct_gain": 5.0,
        "avg_vol": 1000, "sti": 1.5,
    }])
    parent = MainWindow.__new__(MainWindow)
    parent.results_table = ResultsTable()
    parent.results_table.populate(df)
    # User drag → 'sti' moves left of everything else
    parent._results_column_order = [
        "sti", "symbol", "close", "pct_gain", "avg_vol",
    ]
    keys_from_dialog = [
        "sti", "symbol", "close", "pct_gain", "avg_vol",
    ]
    out = parent._build_export_df(df, keys_from_dialog, wants_news=False)
    # Output column ORDER must match the visual order, not canonical.
    assert list(out.columns)[0] == "STI"
    assert list(out.columns)[1] == "Ticker"


# ──────────────────────────────────────────────────────────────────────
# ColumnsManagerDialog — popup widget for reorder + visibility
# ──────────────────────────────────────────────────────────────────────

def test_columns_manager_dialog_drag_emits_columns_updated(_qapp):
    """Drag/drop reorder inside the popup emits `columns_updated`
    with the new ordered_keys + (still-empty) hidden_keys."""
    from PyQt6.QtCore import Qt
    from trade_scanner_fh.gui.dialogs import ColumnsManagerDialog

    cols = [
        ("Ticker", "symbol", str),
        ("Close", "close", str),
        ("Avg Vol", "avg_vol", str),
        ("STI", "sti", str),
    ]
    dlg = ColumnsManagerDialog(
        cols, hidden_keys=set(),
        always_visible_keys={"symbol", "close", "pct_gain"},
    )

    captured: list[tuple] = []
    dlg.columns_updated.connect(
        lambda ordered, hidden: captured.append((list(ordered), list(hidden)))
    )

    # Move row 3 ('sti') to position 0 in the model
    item = dlg._list.takeItem(3)
    dlg._list.insertItem(0, item)
    item.setCheckState(Qt.CheckState.Checked)  # belt-and-suspenders

    # Force the emit pathway directly (Qt model.rowsMoved doesn't
    # always fire for `takeItem`+`insertItem`, but `_emit_state` is
    # the public surface that both paths converge on)
    dlg._emit_state()

    assert captured, "columns_updated was not emitted"
    last_ordered, last_hidden = captured[-1]
    assert last_ordered[0] == "sti"
    assert last_hidden == []


def test_columns_manager_dialog_uncheck_marks_hidden(_qapp):
    """Unchecking a non-core column emits an updated hidden_keys list."""
    from PyQt6.QtCore import Qt
    from trade_scanner_fh.gui.dialogs import ColumnsManagerDialog

    cols = [
        ("Ticker", "symbol", str),
        ("Avg Vol", "avg_vol", str),
        ("STI", "sti", str),
    ]
    dlg = ColumnsManagerDialog(
        cols, hidden_keys=set(),
        always_visible_keys={"symbol"},
    )

    captured: list[tuple] = []
    dlg.columns_updated.connect(
        lambda ordered, hidden: captured.append((list(ordered), list(hidden)))
    )

    # Uncheck Avg Vol (row 1)
    avg_item = dlg._list.item(1)
    assert avg_item.data(Qt.ItemDataRole.UserRole) == "avg_vol"
    avg_item.setCheckState(Qt.CheckState.Unchecked)

    assert captured, "columns_updated was not emitted on uncheck"
    _ordered, hidden = captured[-1]
    assert "avg_vol" in hidden


def test_columns_manager_dialog_blocks_uncheck_of_always_visible(_qapp):
    """Even if a stray Qt event flips an always-visible item to
    Unchecked, the dialog must snap it back."""
    from PyQt6.QtCore import Qt
    from trade_scanner_fh.gui.dialogs import ColumnsManagerDialog

    cols = [
        ("Ticker", "symbol", str),
        ("Close", "close", str),
        ("STI", "sti", str),
    ]
    dlg = ColumnsManagerDialog(
        cols, hidden_keys=set(),
        always_visible_keys={"symbol", "close"},
    )

    sym_item = dlg._list.item(0)
    # The item flag-strip should normally make this a no-op, but
    # exercise the runtime guard anyway by re-enabling the flag.
    sym_item.setFlags(sym_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
    sym_item.setCheckState(Qt.CheckState.Unchecked)
    # Guard re-snaps it to Checked
    assert sym_item.checkState() == Qt.CheckState.Checked


def test_columns_manager_dialog_reset_button_emits_signal(_qapp):
    from trade_scanner_fh.gui.dialogs import ColumnsManagerDialog

    dlg = ColumnsManagerDialog(
        [("Ticker", "symbol", str)],
        hidden_keys=set(),
        always_visible_keys={"symbol"},
    )

    fired: list[bool] = []
    dlg.reset_requested.connect(lambda: fired.append(True))
    dlg._on_reset_clicked()
    assert fired == [True]


# ──────────────────────────────────────────────────────────────────────
# Scan reconcile rule: prepend additions, drop removals
# ──────────────────────────────────────────────────────────────────────

def test_reconcile_prepends_new_columns_to_saved_order(_qapp):
    """A scan that adds NEW columns puts them at the FRONT of the
    saved order. Existing saved-order keys keep their relative
    positions; canonical order of the additions determines their
    relative ordering."""
    from trade_scanner_fh.gui.main_window import MainWindow
    parent = MainWindow.__new__(MainWindow)
    parent._results_column_order = ["sti", "symbol", "close", "pct_gain"]
    canonical = [
        "symbol", "close", "pct_gain",
        "avg_vol", "rs_market", "sti",
    ]
    parent._reconcile_column_order_for_scan(canonical)
    # avg_vol + rs_market are new (canonical order); they prepend
    assert parent._results_column_order[:2] == ["avg_vol", "rs_market"]
    # The user's prior layout follows
    assert parent._results_column_order[2:] == [
        "sti", "symbol", "close", "pct_gain",
    ]


def test_reconcile_drops_removed_columns(_qapp):
    """Variables that no longer produce output columns get dropped
    from the saved order; the rest of the layout survives."""
    from trade_scanner_fh.gui.main_window import MainWindow
    parent = MainWindow.__new__(MainWindow)
    parent._results_column_order = [
        "sti", "symbol", "close", "pct_gain", "avg_vol",
    ]
    # avg_vol no longer in the new scan
    canonical = ["symbol", "close", "pct_gain", "sti"]
    parent._reconcile_column_order_for_scan(canonical)
    assert "avg_vol" not in parent._results_column_order
    # The user's relative order is preserved for surviving columns
    assert parent._results_column_order == [
        "sti", "symbol", "close", "pct_gain",
    ]


def test_reconcile_empty_saved_order_stays_empty(_qapp):
    """When the saved order is empty (default), reconcile is a no-op
    so canonical order applies on the next populate."""
    from trade_scanner_fh.gui.main_window import MainWindow
    parent = MainWindow.__new__(MainWindow)
    parent._results_column_order = []
    parent._reconcile_column_order_for_scan(["a", "b", "c"])
    assert parent._results_column_order == []


# ──────────────────────────────────────────────────────────────────────
# Reset to Default — header right-click + popup button + helper
# ──────────────────────────────────────────────────────────────────────

def test_header_right_click_reset_full_reset_clears_both(_qapp):
    """Header right-click → Reset to Default fires
    `reset_to_default_requested` → ResultsTable forwards via
    `columns_reset_requested` → MainWindow's
    `_reset_columns_to_default` clears BOTH the saved order AND the
    hidden set."""
    from trade_scanner_fh.gui.main_window import MainWindow
    mw = MainWindow()
    try:
        mw._results_column_order = ["sti", "symbol", "close", "pct_gain"]
        mw._deleted_column_keys = {"avg_vol"}
        mw.results_table.columns_reset_requested.emit()
        assert mw._results_column_order == []
        assert mw._deleted_column_keys == set()
    finally:
        mw.close()
        mw.deleteLater()


# ──────────────────────────────────────────────────────────────────────
# Preset save/load: column_order + column_hidden round-trip
# ──────────────────────────────────────────────────────────────────────

def test_preset_save_includes_column_order_and_hidden(_qapp, tmp_path,
                                                      monkeypatch):
    """Saving a preset writes both `column_order` and `column_hidden`."""
    import json
    from trade_scanner_fh.gui import main_window as mw_mod
    monkeypatch.setattr(mw_mod, "PRESETS_DIR", tmp_path)
    mw = mw_mod.MainWindow()
    try:
        mw._results_column_order = ["sti", "symbol", "close"]
        mw._deleted_column_keys = {"avg_vol", "rs_market"}
        mw.preset_combo.addItem("test_layout")
        mw.preset_combo.setCurrentText("test_layout")
        mw._save_preset()
        path = tmp_path / "test_layout.json"
        assert path.exists()
        data = json.loads(path.read_text(encoding="utf-8"))
        # Stays in lockstep with PRESET_SCHEMA_VERSION in main_window.py
        # (bumped to 6 after this assertion was first written; the
        # backward-compat load test below still uses 5 intentionally).
        from trade_scanner_fh.gui.main_window import PRESET_SCHEMA_VERSION
        assert data["_preset_version"] == PRESET_SCHEMA_VERSION
        assert data["column_order"] == ["sti", "symbol", "close"]
        assert sorted(data["column_hidden"]) == ["avg_vol", "rs_market"]
    finally:
        mw.close()
        mw.deleteLater()


def test_preset_load_wipes_results_and_restores_columns(_qapp, tmp_path,
                                                        monkeypatch):
    """Loading a v5 preset wipes `_period_results` + `_active_period`,
    then restores `_results_column_order` and `_deleted_column_keys`."""
    import json
    from trade_scanner_fh.gui import main_window as mw_mod
    monkeypatch.setattr(mw_mod, "PRESETS_DIR", tmp_path)
    # Hand-craft a minimal v5 preset
    preset = {
        "_preset_version": 5,
        "indicators": {},
        "column_order": ["sti", "symbol", "close"],
        "column_hidden": ["avg_vol"],
    }
    (tmp_path / "layout5.json").write_text(
        json.dumps(preset), encoding="utf-8",
    )
    mw = mw_mod.MainWindow()
    try:
        # Pretend a scan has already run
        mw._period_results = {
            "live": pd.DataFrame([{"symbol": "X", "close": 1.0}]),
        }
        mw._period_order = ["live"]
        mw._active_period = "live"

        mw.preset_combo.addItem("layout5")
        mw.preset_combo.setCurrentText("layout5")
        mw._load_preset()

        # Results wiped
        assert mw._period_results == {}
        assert mw._active_period is None
        # Column state restored from preset
        assert mw._results_column_order == ["sti", "symbol", "close"]
        assert mw._deleted_column_keys == {"avg_vol"}
    finally:
        mw.close()
        mw.deleteLater()


def test_preset_load_nonsequenced_reanchors_end_preserves_span(
    _qapp, tmp_path, monkeypatch,
):
    """A NON-sequenced preset re-anchors End to the current most recent
    trading day and preserves the saved window LENGTH for Start (Issue 2):
    a preset saved with a stale 10-day window loads as
    (latest-10d) .. latest, not its stale saved dates."""
    import json
    from datetime import date
    from PyQt6.QtCore import QDate
    from trade_scanner_fh.gui import main_window as mw_mod
    monkeypatch.setattr(mw_mod, "PRESETS_DIR", tmp_path)
    preset = {
        "_preset_version": mw_mod.PRESET_SCHEMA_VERSION,
        "indicators": {},
        "start_date": "2025-01-01",
        "end_date": "2025-01-11",   # 10-day span
        "sequenced_run": False,
    }
    (tmp_path / "stale.json").write_text(json.dumps(preset), encoding="utf-8")
    mw = mw_mod.MainWindow()
    try:
        latest = date(2026, 6, 16)
        mw._latest_data_date = lambda: latest
        mw.preset_combo.addItem("stale")
        mw.preset_combo.setCurrentText("stale")
        assert mw._load_preset() is True

        assert mw.date_end.date() == QDate(2026, 6, 16)
        # Saved span (10 days) preserved, anchored to the fresh End.
        assert mw.date_start.date() == QDate(2026, 6, 6)
    finally:
        mw.close()
        mw.deleteLater()


def test_preset_load_sequenced_keeps_exact_saved_dates(
    _qapp, tmp_path, monkeypatch,
):
    """A sequenced-run preset 'remembers' its exact saved start/end window
    rather than re-anchoring to the latest trading day (Issue 2)."""
    import json
    from datetime import date
    from PyQt6.QtCore import QDate
    from trade_scanner_fh.gui import main_window as mw_mod
    monkeypatch.setattr(mw_mod, "PRESETS_DIR", tmp_path)
    preset = {
        "_preset_version": mw_mod.PRESET_SCHEMA_VERSION,
        "indicators": {},
        "start_date": "2024-03-01",
        "end_date": "2024-09-01",
        "sequenced_run": True,
        "sequenced_cfg": {
            "start": "2024-03-01", "end": "2024-09-01",
            "n": 2, "unit": "months",
        },
    }
    (tmp_path / "seq.json").write_text(json.dumps(preset), encoding="utf-8")
    mw = mw_mod.MainWindow()
    try:
        # Even if 'latest' were consulted, the sequenced branch must ignore it.
        mw._latest_data_date = lambda: date(2026, 6, 16)
        mw.preset_combo.addItem("seq")
        mw.preset_combo.setCurrentText("seq")
        assert mw._load_preset() is True

        assert mw.date_start.date() == QDate(2024, 3, 1)
        assert mw.date_end.date() == QDate(2024, 9, 1)
    finally:
        mw.close()
        mw.deleteLater()


# ──────────────────────────────────────────────────────────────────────
# Toolbar button: opens / focuses the singleton dropdown
# ──────────────────────────────────────────────────────────────────────

def test_columns_button_no_op_pre_scan_no_preset(_qapp):
    """Pre-scan + no preset → click is a no-op (no popup spawned).
    Status bar gets a brief nudge."""
    from trade_scanner_fh.gui.main_window import MainWindow
    mw = MainWindow()
    try:
        # Sanity: no period results, empty saved order
        mw._period_results = {}
        mw._results_column_order = []
        # Click button
        mw._open_columns_dialog()
        assert mw._columns_dialog is None
    finally:
        mw.close()
        mw.deleteLater()


def test_columns_button_spawns_dialog_after_scan(_qapp):
    """After a scan populates the table, clicking Columns ▾ spawns
    the popup with the current active columns listed in visual
    order."""
    from trade_scanner_fh.gui.main_window import MainWindow
    df = pd.DataFrame([{
        "symbol": "A", "close": 100.0, "pct_gain": 5.0, "avg_vol": 1000,
    }])
    mw = MainWindow()
    try:
        mw._period_results = {"live": df}
        mw._period_order = ["live"]
        mw._active_period = "live"
        mw.combo_timeframe.clear()
        mw.combo_timeframe.addItem("live", "live")
        mw.combo_timeframe.setCurrentIndex(0)
        mw._on_timeframe_changed(0)
        mw._open_columns_dialog()
        assert mw._columns_dialog is not None
        keys = mw._columns_dialog.ordered_keys()
        assert "symbol" in keys
        assert "avg_vol" in keys
    finally:
        if mw._columns_dialog is not None:
            mw._columns_dialog.close()
            mw._columns_dialog.deleteLater()
        mw.close()
        mw.deleteLater()
