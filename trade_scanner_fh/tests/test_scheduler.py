"""
F3 — in-app scan scheduler + quick-export tests.

Part A: scheduler core (pure, no window) — entry_is_due time/day/
        enabled/once-per-day logic, ScheduleEntry serialization,
        persistence round-trip + atomicity + corrupt-file tolerance.
Part B: ScanScheduler.check_due glue — fires due entries through the
        preset-load + scan paths, marks last-fired so an entry can't
        fire twice the same day, skips disabled entries, skips +
        logs missing presets, defers (without consuming) while a scan
        is running.
Part C: completion hook — scheduled scans auto quick-export and toast
        via QSystemTrayIcon.showMessage (stub tray); manual scans
        (no pending entry) do neither.
Part D: ExportsController.quick_export — exports-dir + filename shape
        (injected `now`, monkeypatched DATA_DIR), no-results guard,
        preset-name sanitization, multi-period sheets, filesystem
        failure containment (Qt-slot path must never raise).
Part E: unattended-fire hardening — _fire aborts (consumed + logged)
        on preset-load failure / combo mismatch, holds the window's
        _scheduled_fire_active flag around load+run, and
        MainWindow._notify_scan_blocked routes to the log panel (not a
        modal) while the flag is up. Plus source pins for the
        MainWindow→scheduler wiring (on_scan_completed/on_scan_failed
        hook calls + the __init__ scheduler start), mirroring
        test_prefetch_wiring's pin convention.

All time-dependent paths take an injected `now` — no wall-clock
dependence anywhere in this file.
"""

import json
import re
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock

import pandas as pd
import pytest

from trade_scanner_fh import config
from trade_scanner_fh.gui import scheduler as sched_mod
from trade_scanner_fh.gui.scheduler import (
    ScanScheduler, ScheduleEntry, entry_is_due, load_schedules,
    save_schedules,
)

# Fixed fixture dates (no wall clock): 2026-06-08 = Monday(0),
# 2026-06-09 = Tuesday(1), 2026-06-10 = Wednesday(2).
TUE_0930 = datetime(2026, 6, 9, 9, 30)
TUE_0929 = datetime(2026, 6, 9, 9, 29)
TUE_1500 = datetime(2026, 6, 9, 15, 0)
WED_0930 = datetime(2026, 6, 10, 9, 30)


def _entry(**kw) -> ScheduleEntry:
    base = dict(label="Morning", preset_name="Momo", time="09:30",
                days=[1], enabled=True)
    base.update(kw)
    return ScheduleEntry(**base)


# ──────────────────────────────────────────────────────────────────────
# Part A — pure core
# ──────────────────────────────────────────────────────────────────────

def test_due_when_time_and_day_match():
    e = _entry()
    assert entry_is_due(e, TUE_0930) is True          # exact time
    assert entry_is_due(e, TUE_1500) is True          # later same day
    assert entry_is_due(e, TUE_0929) is False         # a minute early
    assert entry_is_due(e, WED_0930) is False         # wrong weekday


def test_not_due_when_disabled():
    assert entry_is_due(_entry(enabled=False), TUE_0930) is False


def test_fires_at_most_once_per_day():
    e = _entry(last_fired_date="2026-06-09")
    assert entry_is_due(e, TUE_1500) is False
    # ...but a prior-day marker doesn't block today
    e2 = _entry(days=[1, 2], last_fired_date="2026-06-09")
    assert entry_is_due(e2, WED_0930) is True


def test_malformed_time_or_empty_days_never_due():
    assert entry_is_due(_entry(time="garbage"), TUE_0930) is False
    assert entry_is_due(_entry(time="25:99"), TUE_0930) is False
    assert entry_is_due(_entry(days=[]), TUE_0930) is False


def test_from_dict_tolerates_missing_and_bad_keys():
    e = ScheduleEntry.from_dict({})
    assert e.label == "" and e.preset_name == ""
    assert e.enabled is True and e.last_fired_date == ""
    assert e.days == []
    # Day values are coerced, deduped, and clamped to 0–6
    e2 = ScheduleEntry.from_dict(
        {"days": [1, "2", 2, 99, -1, "x", None], "enabled": 0}
    )
    assert e2.days == [1, 2]
    assert e2.enabled is False


def _sched_path(tmp_path) -> Path:
    return tmp_path / "schedules.json"


def test_persistence_roundtrip(tmp_path):
    p = _sched_path(tmp_path)
    entries = [
        _entry(),
        _entry(label="Lunch", time="12:15", days=[0, 2, 4],
               enabled=False, last_fired_date="2026-06-08"),
    ]
    save_schedules(entries, path=p)
    loaded = load_schedules(path=p)
    assert loaded == entries
    data = json.loads(p.read_text(encoding="utf-8"))
    assert data["version"] == sched_mod.SCHEDULES_VERSION
    assert len(data["entries"]) == 2


def test_save_goes_through_atomic_write(tmp_path, monkeypatch):
    p = _sched_path(tmp_path)
    calls = []
    orig = config.atomic_write_text

    def spy(path, content, encoding="utf-8"):
        calls.append(Path(path))
        orig(path, content, encoding)

    monkeypatch.setattr(config, "atomic_write_text", spy)
    save_schedules([_entry()], path=p)
    assert calls == [p], "save_schedules must write via atomic_write_text"


def test_corrupt_or_wrong_shape_file_loads_as_empty(tmp_path):
    p = _sched_path(tmp_path)
    p.write_text("{not valid json!!", encoding="utf-8")
    assert load_schedules(path=p) == []
    p.write_text(json.dumps([1, 2, 3]), encoding="utf-8")
    assert load_schedules(path=p) == []
    p.write_text(json.dumps({"entries": "nope"}), encoding="utf-8")
    assert load_schedules(path=p) == []
    # Non-dict entries are skipped, dict entries survive
    p.write_text(
        json.dumps({"version": 1,
                    "entries": [42, {"label": "ok", "preset_name": "P"}]}),
        encoding="utf-8",
    )
    loaded = load_schedules(path=p)
    assert len(loaded) == 1 and loaded[0].label == "ok"
    # Missing file
    assert load_schedules(path=tmp_path / "absent.json") == []


def test_default_path_honors_monkeypatched_data_dir(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    save_schedules([_entry()])
    assert (tmp_path / "schedules.json").exists()
    assert load_schedules()[0].label == "Morning"


# ──────────────────────────────────────────────────────────────────────
# Part B — ScanScheduler.check_due glue
# ──────────────────────────────────────────────────────────────────────

class _StubWin:
    """Duck-typed MainWindow stand-in: just the delegates check_due /
    _fire / on_scan_completed touch."""

    def __init__(self):
        self._worker = None
        self.log_lines: list[str] = []
        self.log_panel = SimpleNamespace(write_line=self.log_lines.append)
        self.preset_combo = MagicMock()
        self.preset_combo.findText.return_value = 1
        self._refresh_preset_list = MagicMock()
        self._load_preset = MagicMock()
        self._run_scan = MagicMock(side_effect=self._start_scan)
        self._quick_export = MagicMock(return_value=None)
        self._period_order: list[str] = []
        self._period_results: dict[str, pd.DataFrame] = {}

    def _start_scan(self):
        self._worker = object()   # simulate a started ScanWorker

    def finish_scan(self, sched: ScanScheduler):
        self._worker = None
        sched.on_scan_completed()


@pytest.fixture
def world(tmp_path, monkeypatch, _qapp):
    """Monkeypatched DATA_DIR + a saved 'Momo' preset on disk + a
    scheduler wired to a stub window with an injected stub tray."""
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    presets = tmp_path / "presets"
    presets.mkdir()
    (presets / "Momo.json").write_text("{}", encoding="utf-8")
    win = _StubWin()
    sched = ScanScheduler(win)
    sched._tray = MagicMock()     # stub tray — no real QSystemTrayIcon
    return SimpleNamespace(tmp_path=tmp_path, win=win, sched=sched)


def test_due_entry_fires_through_preset_and_scan_paths(world):
    win, sched = world.win, world.sched
    sched.entries = [_entry()]
    sched.check_due(TUE_0930)

    win.preset_combo.setCurrentIndex.assert_called_once_with(1)
    win._load_preset.assert_called_once()
    win._run_scan.assert_called_once()
    assert sched._pending is sched.entries[0]
    assert sched.entries[0].last_fired_date == "2026-06-09"
    # Fired marker persisted so a relaunch the same day can't re-fire
    data = json.loads(
        (world.tmp_path / "schedules.json").read_text(encoding="utf-8"))
    assert data["entries"][0]["last_fired_date"] == "2026-06-09"


def test_entry_does_not_fire_twice_same_day(world):
    win, sched = world.win, world.sched
    sched.entries = [_entry()]
    sched.check_due(TUE_0930)
    win.finish_scan(sched)        # scan completes, scheduler idle again
    sched.check_due(TUE_1500)     # later the same day
    assert win._run_scan.call_count == 1
    sched.check_due(WED_0930)     # Wednesday isn't in days=[Tue]
    assert win._run_scan.call_count == 1


def test_entry_fires_again_next_configured_day(world):
    win, sched = world.win, world.sched
    sched.entries = [_entry(days=[1, 2], last_fired_date="2026-06-09")]
    sched.check_due(WED_0930)
    assert win._run_scan.call_count == 1
    assert sched.entries[0].last_fired_date == "2026-06-10"


def test_disabled_entry_is_skipped(world):
    win, sched = world.win, world.sched
    sched.entries = [_entry(enabled=False)]
    sched.check_due(TUE_0930)
    win._run_scan.assert_not_called()
    win._load_preset.assert_not_called()
    assert sched.entries[0].last_fired_date == ""
    assert win.log_lines == []


def test_missing_preset_skipped_logged_and_consumed(world):
    win, sched = world.win, world.sched
    sched.entries = [_entry(label="Ghost run", preset_name="Deleted")]
    sched.check_due(TUE_0930)
    win._run_scan.assert_not_called()
    win._load_preset.assert_not_called()
    assert any(
        "Ghost run" in line and "no longer exists" in line
        for line in win.log_lines
    )
    # Consumed for the day — the next tick doesn't re-log all day long
    assert sched.entries[0].last_fired_date == "2026-06-09"
    n_lines = len(win.log_lines)
    sched.check_due(TUE_1500)
    assert len(win.log_lines) == n_lines


def test_busy_scanner_defers_without_consuming_then_fires(world):
    win, sched = world.win, world.sched
    sched.entries = [_entry()]
    win._worker = object()        # a manual scan is running
    sched.check_due(TUE_0930)
    win._run_scan.assert_not_called()
    assert sched.entries[0].last_fired_date == ""
    assert any("already running" in line for line in win.log_lines)
    # The deferral log line isn't repeated on every 30s tick
    n_lines = len(win.log_lines)
    sched.check_due(TUE_0930)
    assert len(win.log_lines) == n_lines
    # Manual scan finishes → next tick fires the entry
    win._worker = None
    sched.on_scan_completed()     # manual completion: no toast, clears throttle
    sched.check_due(TUE_1500)
    win._run_scan.assert_called_once()
    assert sched.entries[0].last_fired_date == "2026-06-09"


def test_only_one_launch_per_tick_second_entry_fires_later(world):
    win, sched = world.win, world.sched
    sched.entries = [_entry(label="A"), _entry(label="B")]
    sched.check_due(TUE_0930)
    assert win._run_scan.call_count == 1
    assert sched.entries[0].last_fired_date == "2026-06-09"
    assert sched.entries[1].last_fired_date == ""
    win.finish_scan(sched)
    sched.check_due(TUE_1500)
    assert win._run_scan.call_count == 2
    assert sched.entries[1].last_fired_date == "2026-06-09"


def test_declined_scan_clears_pending_so_future_fires_work(world):
    win, sched = world.win, world.sched
    win._run_scan = MagicMock()   # declines: leaves win._worker = None
    sched.entries = [_entry(days=[1, 2])]
    sched.check_due(TUE_0930)
    assert sched._pending is None
    assert any("did not start" in line for line in win.log_lines)
    # Next configured day still fires
    sched.check_due(WED_0930)
    assert win._run_scan.call_count == 2


def test_scan_failed_hook_clears_pending(world):
    win, sched = world.win, world.sched
    sched.entries = [_entry()]
    sched.check_due(TUE_0930)
    assert sched._pending is not None
    win._worker = None
    sched.on_scan_failed()
    assert sched._pending is None
    assert any("failed" in line for line in win.log_lines)
    sched._tray.showMessage.assert_not_called()


def test_fire_aborts_when_preset_load_fails(world):
    """_load_preset returning False (missing/corrupt preset) must abort
    the fire — no scan under whatever settings happen to be active, no
    pending entry — while the entry stays consumed for the day."""
    win, sched = world.win, world.sched
    win._load_preset = MagicMock(return_value=False)
    sched.entries = [_entry()]
    sched.check_due(TUE_0930)
    win._run_scan.assert_not_called()
    assert sched._pending is None
    assert any("failed to load" in line for line in win.log_lines)
    assert sched.entries[0].last_fired_date == "2026-06-09"
    # finally-block hygiene: the modal-suppression flag is back down
    assert win._scheduled_fire_active is False


def test_fire_aborts_when_combo_select_misses(world):
    """If the combo doesn't end up showing the scheduled preset (rename
    race / select failure), firing anyway would scan + diff + export
    under the WRONG settings — abort instead (consumed + logged)."""
    win, sched = world.win, world.sched
    win.preset_combo.currentText.return_value = "SomeOtherPreset"
    sched.entries = [_entry()]
    sched.check_due(TUE_0930)
    win._load_preset.assert_not_called()
    win._run_scan.assert_not_called()
    assert sched._pending is None
    assert any("could not select preset" in line for line in win.log_lines)
    assert sched.entries[0].last_fired_date == "2026-06-09"


def test_fire_holds_scheduled_flag_during_load_and_run(world):
    """_scheduled_fire_active must be True while _load_preset and
    _run_scan execute (so their 'cannot start' notices log instead of
    raising modals on an unattended machine) and False afterwards."""
    win, sched = world.win, world.sched
    seen: list[tuple[str, bool]] = []
    win._load_preset = MagicMock(side_effect=lambda: (
        seen.append(("load", win._scheduled_fire_active)) or True))

    def _run():
        seen.append(("run", win._scheduled_fire_active))
        win._worker = object()

    win._run_scan = MagicMock(side_effect=_run)
    sched.entries = [_entry()]
    sched.check_due(TUE_0930)
    assert seen == [("load", True), ("run", True)]
    assert win._scheduled_fire_active is False


# ──────────────────────────────────────────────────────────────────────
# Part C — completion hook: quick-export + toast
# ──────────────────────────────────────────────────────────────────────

def test_scheduled_completion_quick_exports_and_toasts(world):
    win, sched = world.win, world.sched
    sched.entries = [_entry(label="Morning")]
    sched.check_due(TUE_0930)

    # Scan-done state the window would hold after the F2 diff ran.
    df = pd.DataFrame({
        "symbol": ["AAA", "BBB", "CCC"],
        "chg": ["NEW", "", ""],
    })
    win._period_order = ["1D"]
    win._period_results = {"1D": df}
    win._quick_export = MagicMock(
        return_value=r"C:\fake\exports\scan_Momo_20260609-093001.xlsx")

    win.finish_scan(sched)

    win._quick_export.assert_called_once()
    sched._tray.showMessage.assert_called_once()
    args = sched._tray.showMessage.call_args[0]
    joined = " ".join(str(a) for a in args[:2])
    assert "Scheduled scan 'Morning': 3 results, 1 new" in joined
    # Log line carries the export path
    assert any("scan_Momo_20260609-093001.xlsx" in line
               for line in win.log_lines)
    assert sched._pending is None


def test_manual_scan_completion_is_a_noop(world):
    win, sched = world.win, world.sched
    assert sched._pending is None
    sched.on_scan_completed()
    win._quick_export.assert_not_called()
    sched._tray.showMessage.assert_not_called()


# ──────────────────────────────────────────────────────────────────────
# Part D — quick-export filename / dir shape
# ──────────────────────────────────────────────────────────────────────

def _export_shell(tmp_path, monkeypatch, period_results,
                  preset_text="(select preset)"):
    """Bare MainWindow shell carrying exactly what quick_export and the
    real `_write_xlsx_multi_sheet` pipeline touch."""
    from trade_scanner_fh.gui.main_window import MainWindow
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    win = MainWindow.__new__(MainWindow)
    win._period_results = period_results
    win._period_order = list(period_results.keys())
    win._active_period = win._period_order[0] if win._period_order else None
    win._results_column_order = []
    win._apply_view_filters = lambda df: df
    win.preset_combo = MagicMock()
    win.preset_combo.currentText.return_value = preset_text
    win.results_table = MagicMock()
    win.results_table.active_columns = [
        ("Ticker", "symbol", str),
        ("Close", "close", lambda x: f"${x:.2f}"),
    ]
    win.log_panel = MagicMock()
    win._log_lines = []
    win.log_panel.write_line.side_effect = win._log_lines.append
    win.status = MagicMock()
    return win


def _df(*syms):
    return pd.DataFrame(
        {"symbol": list(syms), "close": [10.0] * len(syms)})


def test_quick_export_writes_timestamped_xlsx(_qapp, tmp_path, monkeypatch):
    win = _export_shell(tmp_path, monkeypatch, {"1D": _df("AAA", "BBB")})
    out = win._quick_export(now=datetime(2026, 6, 11, 9, 30, 5))

    expect = tmp_path / "exports" / "scan_adhoc_20260611-093005.xlsx"
    assert out == str(expect)
    assert expect.exists()
    assert re.fullmatch(r"scan_adhoc_\d{8}-\d{6}\.xlsx", expect.name)
    # Log-panel line carries the full path
    assert any(str(expect) in line for line in win._log_lines)

    # Content sanity via the same engine the exporter used
    from openpyxl import load_workbook
    wb = load_workbook(expect)
    assert wb.sheetnames == ["1D"]
    ws = wb["1D"]
    assert [c.value for c in ws[1]] == ["Ticker", "Close"]
    assert ws.max_row == 3   # header + 2 data rows


def test_quick_export_uses_preset_key_in_filename(_qapp, tmp_path,
                                                  monkeypatch):
    win = _export_shell(tmp_path, monkeypatch, {"1D": _df("AAA")},
                        preset_text="Momo Setup")
    out = win._quick_export(now=datetime(2026, 6, 11, 16, 0, 0))
    assert Path(out).name == "scan_Momo Setup_20260611-160000.xlsx"
    assert Path(out).parent == tmp_path / "exports"


def test_quick_export_sanitizes_illegal_filename_chars():
    from trade_scanner_fh.gui.exports import ExportsController
    f = ExportsController._safe_filename_component
    assert f("Bad/Name") == "Bad_Name"
    assert f('a<b>:"c"|?*') == "a_b___c____"
    assert f("") == "adhoc"
    assert f("  . ") == "adhoc"
    assert f("clean-name_1") == "clean-name_1"


def test_quick_export_multi_period_one_sheet_each(_qapp, tmp_path,
                                                  monkeypatch):
    win = _export_shell(tmp_path, monkeypatch, {
        "1D": _df("AAA"),
        "1W": _df("AAA", "BBB", "CCC"),
    })
    out = win._quick_export(now=datetime(2026, 6, 11, 9, 0, 0))
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert wb.sheetnames == ["1D", "1W"]
    assert wb["1W"].max_row == 4


def test_quick_export_no_results_logs_and_creates_nothing(
        _qapp, tmp_path, monkeypatch):
    win = _export_shell(tmp_path, monkeypatch, {})
    out = win._quick_export(now=datetime(2026, 6, 11, 9, 0, 0))
    assert out is None
    # Exports dir is created lazily — not on a no-op
    assert not (tmp_path / "exports").exists()
    assert any("no results" in line for line in win._log_lines)


def test_quick_export_mkdir_failure_logs_and_returns_none(
        _qapp, tmp_path, monkeypatch):
    """quick_export runs in a Qt-slot path (Scans → Quick Export menu):
    a filesystem failure in the exports-dir mkdir must be contained to
    a logged None return, never an exception escaping the slot (the
    PyInstaller-windowed-mode abort this codebase guards against)."""
    win = _export_shell(tmp_path, monkeypatch, {"1D": _df("AAA")})
    # Squat a FILE on the exports dir name → mkdir(parents=True,
    # exist_ok=True) raises FileExistsError (an OSError subclass).
    (tmp_path / "exports").write_text("not a directory", encoding="utf-8")
    out = win._quick_export(now=datetime(2026, 6, 11, 9, 0, 0))
    assert out is None
    assert any("Quick Export failed" in line for line in win._log_lines)


# ──────────────────────────────────────────────────────────────────────
# Part E — unattended-fire modal suppression + MainWindow wiring pins
# ──────────────────────────────────────────────────────────────────────

def _notify_shell(_qapp):
    """Bare MainWindow shell carrying just what _notify_scan_blocked
    touches (log panel + the scheduled-fire flag)."""
    from trade_scanner_fh.gui.main_window import MainWindow
    win = MainWindow.__new__(MainWindow)
    win._log_lines = []
    win.log_panel = SimpleNamespace(write_line=win._log_lines.append)
    return win


def test_notify_scan_blocked_logs_during_scheduled_fire(_qapp, monkeypatch):
    from trade_scanner_fh.gui import main_window as mw_mod
    calls: list[tuple] = []
    monkeypatch.setattr(mw_mod.QMessageBox, "warning",
                        lambda *a, **k: calls.append(("warning", a)))
    monkeypatch.setattr(mw_mod.QMessageBox, "information",
                        lambda *a, **k: calls.append(("info", a)))
    win = _notify_shell(_qapp)
    win._scheduled_fire_active = True
    mw_mod.MainWindow._notify_scan_blocked(win, "No Data", "no cache yet")
    mw_mod.MainWindow._notify_scan_blocked(
        win, "All Omitted", "all seen", icon="info")
    assert calls == [], "scheduled fires must never raise a modal"
    assert win._log_lines == [
        "No Data: no cache yet", "All Omitted: all seen",
    ]


def test_notify_scan_blocked_is_modal_interactively(_qapp, monkeypatch):
    from trade_scanner_fh.gui import main_window as mw_mod
    calls: list[tuple] = []
    monkeypatch.setattr(mw_mod.QMessageBox, "warning",
                        lambda *a, **k: calls.append(("warning", a)))
    monkeypatch.setattr(mw_mod.QMessageBox, "information",
                        lambda *a, **k: calls.append(("info", a)))
    win = _notify_shell(_qapp)
    # Flag absent (bare shell) and explicitly False both mean modal.
    mw_mod.MainWindow._notify_scan_blocked(win, "No Data", "msg")
    win._scheduled_fire_active = False
    mw_mod.MainWindow._notify_scan_blocked(win, "All Omitted", "msg",
                                           icon="info")
    assert [kind for kind, _a in calls] == ["warning", "info"]
    assert win._log_lines == []


# Source pins — MainWindow→scheduler wiring (same convention as
# test_prefetch_wiring.py: the hooks are reached via __dict__ lookups /
# dynamic dispatch the type checker and duck-typed Part B/C stubs can't
# see, so renaming or dropping a call would otherwise stay green).

def _main_window_source() -> str:
    import trade_scanner_fh.gui.main_window as mw
    return open(mw.__file__, encoding="utf-8").read()


def _method_body(text: str, name: str) -> str:
    start = text.index(f"    def {name}(")
    ends = [i for i in (text.find("\n    def ", start + 1),
                        text.find("\ndef ", start + 1)) if i != -1]
    return text[start:min(ends)] if ends else text[start:]


def test_pin_scan_done_impl_notifies_scheduler():
    """The completion hook (quick-export + toast + pending-entry clear)
    hangs off _on_scan_done_impl's on_scan_completed call — deleting it
    would silently kill the whole F3 completion flow."""
    body = _method_body(_main_window_source(), "_on_scan_done_impl")
    assert "on_scan_completed" in body


def test_pin_scan_crash_handler_notifies_scheduler():
    """The crash path must call on_scan_failed or a crashed scheduled
    scan leaves _pending dangling and blocks every future fire."""
    body = _method_body(_main_window_source(), "_on_scan_done")
    assert "on_scan_failed" in body


def test_pin_init_starts_scheduler():
    """Fully-constructed windows must start the ~30s due-check timer."""
    body = _method_body(_main_window_source(), "__init__")
    assert "self._scheduler.start()" in body


def test_pin_no_direct_modals_in_scan_start_paths():
    """_run_scan and _load_preset must route every 'cannot start'
    notice through _notify_scan_blocked — a raw QMessageBox reintroduces
    the unattended-machine modal hang for scheduled fires."""
    text = _main_window_source()
    for name in ("_run_scan", "_load_preset"):
        body = _method_body(text, name)
        assert "QMessageBox" not in body, name
        assert "_notify_scan_blocked" in body, name
